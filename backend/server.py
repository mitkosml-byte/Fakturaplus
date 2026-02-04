from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Response, Request
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
from enum import Enum
import uuid
from datetime import datetime, timezone, timedelta
import base64
import httpx
import io
import re
from passlib.context import CryptContext

# Rate limiting
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Environment
IS_PRODUCTION = os.environ.get('ENVIRONMENT', 'development') == 'production'

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Rate limiter
limiter = Limiter(key_func=get_remote_address)

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'test_database')]

# Emergent LLM Key
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')

# Create the main app
app = FastAPI(
    title="Invoice Manager API",
    version="1.0.0",
    docs_url=None if IS_PRODUCTION else "/docs",
    redoc_url=None if IS_PRODUCTION else "/redoc"
)

# Add rate limiter
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging - less verbose in production
log_level = logging.WARNING if IS_PRODUCTION else logging.INFO
logging.basicConfig(
    level=log_level,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ===================== MODELS =====================

class Company(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # Име на фирмата
    eik: str  # ЕИК/Булстат
    vat_number: Optional[str] = None  # ДДС номер (BG + ЕИК)
    mol: Optional[str] = None  # МОЛ (Материално отговорно лице)
    address: Optional[str] = None  # Адрес
    city: Optional[str] = None  # Град
    phone: Optional[str] = None  # Телефон
    email: Optional[str] = None  # Имейл
    bank_name: Optional[str] = None  # Банка
    bank_iban: Optional[str] = None  # IBAN
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CompanyCreate(BaseModel):
    name: str
    eik: str
    vat_number: Optional[str] = None
    mol: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    bank_name: Optional[str] = None
    bank_iban: Optional[str] = None

class CompanyUpdate(BaseModel):
    name: Optional[str] = None
    eik: Optional[str] = None
    vat_number: Optional[str] = None
    mol: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    bank_name: Optional[str] = None
    bank_iban: Optional[str] = None

class User(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    role: str = "staff"  # "owner", "manager", or "staff"
    company_id: Optional[str] = None  # Връзка към фирмата
    password_hash: Optional[str] = None  # За email/password auth
    auth_provider: str = "email"  # "google" or "email"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserSession(BaseModel):
    user_id: str
    session_token: str
    expires_at: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Email/Password Auth Models
class UserRegister(BaseModel):
    email: str
    password: str
    name: str

class UserLogin(BaseModel):
    email: str
    password: str

# Invitation model for user invitations
class Invitation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    invited_by: str  # user_id на изпращача
    email: Optional[str] = None
    phone: Optional[str] = None
    role: str = "staff"  # Роля за поканения
    code: str = Field(default_factory=lambda: uuid.uuid4().hex[:8].upper())  # 8-символен код
    status: str = "pending"  # "pending", "accepted", "cancelled", "expired"
    expires_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc) + timedelta(days=7))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InvitationCreate(BaseModel):
    email: Optional[str] = None
    phone: Optional[str] = None
    role: str = "staff"

class Invoice(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    company_id: Optional[str] = None  # Връзка към фирмата
    supplier: str
    invoice_number: str
    amount_without_vat: float
    vat_amount: float
    total_amount: float
    date: datetime
    image_base64: Optional[str] = None
    notes: Optional[str] = None
    items: Optional[List[dict]] = None  # Списък с артикули
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Invoice Item models
class InvoiceItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # Име на артикула
    quantity: float = 1  # Количество
    unit: str = "бр."  # Мерна единица (бр., кг., л., м.)
    unit_price: float  # Единична цена без ДДС
    total_price: float  # Обща цена без ДДС
    vat_amount: float = 0  # ДДС за артикула

class InvoiceItemCreate(BaseModel):
    name: str
    quantity: float = 1
    unit: str = "бр."
    unit_price: float
    total_price: Optional[float] = None  # Ако не е подадено, се изчислява
    vat_amount: Optional[float] = None

# Item Price History - за проследяване на цените
class ItemPriceHistory(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    supplier: str  # Доставчик
    item_name: str  # Нормализирано име на артикул
    unit_price: float  # Единична цена
    quantity: float  # Количество
    unit: str  # Мерна единица
    invoice_id: str  # Връзка към фактурата
    invoice_number: str
    invoice_date: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Price Alert settings
class PriceAlertSettings(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    threshold_percent: float = 10.0  # Праг за аларма в %
    enabled: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PriceAlert(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    item_name: str
    supplier: str
    old_price: float
    new_price: float
    change_percent: float
    invoice_id: str
    invoice_number: str
    status: str = "unread"  # unread, read, dismissed
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InvoiceCreate(BaseModel):
    supplier: str
    invoice_number: str
    amount_without_vat: float
    vat_amount: float
    total_amount: float
    date: str
    image_base64: Optional[str] = None
    notes: Optional[str] = None
    items: Optional[List[InvoiceItemCreate]] = None  # Артикули

class InvoiceUpdate(BaseModel):
    supplier: Optional[str] = None
    invoice_number: Optional[str] = None
    amount_without_vat: Optional[float] = None
    vat_amount: Optional[float] = None
    total_amount: Optional[float] = None
    date: Optional[str] = None
    notes: Optional[str] = None

class DailyRevenue(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    date: str
    fiscal_revenue: float = 0
    pocket_money: float = 0  # "джобче" - не влиза в ДДС
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DailyRevenueCreate(BaseModel):
    date: str
    fiscal_revenue: float = 0
    pocket_money: float = 0

class NonInvoiceExpense(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    description: str
    amount: float
    date: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NonInvoiceExpenseCreate(BaseModel):
    description: str
    amount: float
    date: str

# ===================== PERSONAL EXPENSES & ROI MODELS =====================

class PersonalExpenseType(str, Enum):
    INVESTMENT = "investment"  # Инвестиция
    RECURRING = "recurring"    # Текущ разход
    ONE_TIME = "one_time"      # Еднократен

class PersonalExpenseCategory(str, Enum):
    GOODS = "goods"            # Стока
    SERVICE = "service"        # Услуга
    PERSONNEL = "personnel"    # Персонал
    RENT = "rent"              # Наем
    EXTRAORDINARY = "extraordinary"  # Извънреден
    OTHER = "other"            # Друго

class PersonalExpense(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    company_id: str
    amount: float
    description: str
    expense_type: PersonalExpenseType = PersonalExpenseType.RECURRING
    category: PersonalExpenseCategory = PersonalExpenseCategory.OTHER
    period_month: int  # 1-12
    period_year: int
    supplier_id: Optional[str] = None  # Връзка с доставчик (ако има)
    project_name: Optional[str] = None  # Връзка с проект (ако има)
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PersonalExpenseCreate(BaseModel):
    amount: float
    description: str
    expense_type: str = "recurring"
    category: str = "other"
    period_month: int
    period_year: int
    supplier_id: Optional[str] = None
    project_name: Optional[str] = None
    notes: Optional[str] = None

class ROIAnalysis(BaseModel):
    period_start: str
    period_end: str
    total_personal_investment: float
    total_revenue: float
    total_profit: float
    roi_percent: float
    is_profitable: bool
    investment_covered: bool
    ai_insights: List[str]

class NotificationSettings(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    vat_threshold_enabled: bool = False
    vat_threshold_amount: float = 0
    periodic_enabled: bool = False
    periodic_dates: List[int] = []  # Days of month (1-31)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NotificationSettingsUpdate(BaseModel):
    vat_threshold_enabled: Optional[bool] = None
    vat_threshold_amount: Optional[float] = None
    periodic_enabled: Optional[bool] = None
    periodic_dates: Optional[List[int]] = None

class OCRResult(BaseModel):
    supplier: str
    invoice_number: str
    amount_without_vat: float
    vat_amount: float
    total_amount: float
    invoice_date: Optional[str] = None  # Дата на издаване от фактурата
    corrections: Optional[List[str]] = None  # Списък с направени корекции
    confidence: Optional[float] = None  # Увереност в резултата (0-1)

class SessionDataResponse(BaseModel):
    id: str
    email: str
    name: str
    picture: Optional[str] = None
    session_token: str

# ===================== AUTH HELPERS =====================

async def get_session_token(request: Request) -> Optional[str]:
    # Check cookie first
    session_token = request.cookies.get("session_token")
    if session_token:
        return session_token
    # Check Authorization header
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.startswith("Bearer "):
        return auth_header.split(" ")[1]
    return None

async def get_current_user(request: Request) -> User:
    session_token = await get_session_token(request)
    if not session_token:
        raise HTTPException(status_code=401, detail="Не сте влезли в системата")
    
    session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Невалидна сесия")
    
    expires_at = session["expires_at"]
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Сесията е изтекла")
    
    user_doc = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Потребителят не е намерен")
    
    return User(**user_doc)

async def get_current_user_optional(request: Request) -> Optional[User]:
    try:
        return await get_current_user(request)
    except Exception:
        return None

# ===================== AUTH ENDPOINTS =====================

@api_router.post("/auth/session")
async def create_session(request: Request, response: Response):
    body = await request.json()
    session_id = body.get("session_id")
    
    if not session_id:
        raise HTTPException(status_code=400, detail="Липсва session_id")
    
    # Exchange session_id for user data
    async with httpx.AsyncClient() as client_http:
        resp = await client_http.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": session_id}
        )
        if resp.status_code != 200:
            raise HTTPException(status_code=401, detail="Невалиден session_id")
        user_data = resp.json()
    
    session_data = SessionDataResponse(**user_data)
    
    # Check if user exists
    existing_user = await db.users.find_one({"email": session_data.email}, {"_id": 0})
    
    if not existing_user:
        # Create new user
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        
        # Auto-create company for new user
        company_name = session_data.name.split()[0] + " Company" if session_data.name else "My Company"
        new_company = Company(
            name=company_name,
            eik=f"AUTO{uuid.uuid4().hex[:9].upper()}"  # Temporary auto-generated EIK
        )
        await db.companies.insert_one(new_company.dict())
        
        new_user = {
            "user_id": user_id,
            "email": session_data.email,
            "name": session_data.name,
            "picture": session_data.picture,
            "role": "owner",  # First user is owner
            "company_id": new_company.id,
            "auth_provider": "google",
            "created_at": datetime.now(timezone.utc)
        }
        await db.users.insert_one(new_user)
    else:
        user_id = existing_user["user_id"]
    
    # Create session
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    session_doc = {
        "user_id": user_id,
        "session_token": session_data.session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc)
    }
    await db.user_sessions.insert_one(session_doc)
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_data.session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7 * 24 * 60 * 60,
        path="/"
    )
    
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return {"user": user_doc, "session_token": session_data.session_token}

@api_router.get("/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user.dict()

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    session_token = await get_session_token(request)
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    response.delete_cookie(key="session_token", path="/")
    return {"message": "Успешно излязохте"}

# ===================== EMAIL/PASSWORD AUTH =====================

def validate_email(email: str) -> bool:
    """Validate email format"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def validate_password(password: str) -> tuple[bool, str]:
    """Validate password strength"""
    if len(password) < 8:
        return False, "Паролата трябва да е поне 8 символа"
    if not re.search(r'[A-Za-z]', password):
        return False, "Паролата трябва да съдържа поне една буква"
    if not re.search(r'\d', password):
        return False, "Паролата трябва да съдържа поне една цифра"
    return True, ""

@api_router.post("/auth/register")
async def register_user(user_data: UserRegister, response: Response):
    """Register new user with email/password"""
    # Validate email
    if not validate_email(user_data.email):
        raise HTTPException(status_code=400, detail="Невалиден имейл адрес")
    
    # Validate password
    is_valid, error_msg = validate_password(user_data.password)
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Validate name
    if not user_data.name or len(user_data.name.strip()) < 2:
        raise HTTPException(status_code=400, detail="Името трябва да е поне 2 символа")
    
    # Check if user already exists
    existing_user = await db.users.find_one({"email": user_data.email.lower()})
    if existing_user:
        raise HTTPException(status_code=400, detail="Потребител с този имейл вече съществува")
    
    # Create user
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    password_hash = pwd_context.hash(user_data.password)
    
    # Auto-create company for new user
    company_name = user_data.name.split()[0] + " Company" if user_data.name else "My Company"
    new_company = Company(
        name=company_name,
        eik=f"AUTO{uuid.uuid4().hex[:9].upper()}"
    )
    await db.companies.insert_one(new_company.dict())
    
    new_user = {
        "user_id": user_id,
        "email": user_data.email.lower(),
        "name": user_data.name.strip(),
        "picture": None,
        "role": "owner",
        "company_id": new_company.id,
        "password_hash": password_hash,
        "auth_provider": "email",
        "created_at": datetime.now(timezone.utc)
    }
    await db.users.insert_one(new_user)
    
    # Create session
    session_token = uuid.uuid4().hex
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    session_doc = {
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc)
    }
    await db.user_sessions.insert_one(session_doc)
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7 * 24 * 60 * 60,
        path="/"
    )
    
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    return {"user": user_doc, "session_token": session_token}

@api_router.post("/auth/login")
async def login_user(user_data: UserLogin, response: Response):
    """Login with email/password"""
    # Find user
    user = await db.users.find_one({"email": user_data.email.lower()})
    if not user:
        raise HTTPException(status_code=401, detail="Невалиден имейл или парола")
    
    # Check if user has password (might be Google-only user)
    if not user.get("password_hash"):
        raise HTTPException(status_code=401, detail="Този акаунт използва Google вход. Моля, използвайте бутона за Google.")
    
    # Verify password
    if not pwd_context.verify(user_data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Невалиден имейл или парола")
    
    # Create session
    session_token = uuid.uuid4().hex
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    session_doc = {
        "user_id": user["user_id"],
        "session_token": session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc)
    }
    await db.user_sessions.insert_one(session_doc)
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7 * 24 * 60 * 60,
        path="/"
    )
    
    user_doc = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0, "password_hash": 0})
    return {"user": user_doc, "session_token": session_token}

@api_router.put("/auth/role/{user_id}")
async def update_user_role(user_id: str, request: Request, current_user: User = Depends(get_current_user)):
    body = await request.json()
    role = body.get("role")
    
    if current_user.role != "owner":
        raise HTTPException(status_code=403, detail="Само титулярят може да променя роли")
    
    if role not in ["owner", "manager", "staff"]:
        raise HTTPException(status_code=400, detail="Невалидна роля. Допустими: owner, manager, staff")
    
    # Get target user
    target_user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="Потребителят не е намерен")
    
    # Ensure same company
    if target_user.get("company_id") != current_user.company_id:
        raise HTTPException(status_code=403, detail="Потребителят не е от вашата фирма")
    
    # Cannot change own role if owner
    if user_id == current_user.user_id and current_user.role == "owner":
        raise HTTPException(status_code=400, detail="Не можете да променяте собствената си роля на собственик")
    
    result = await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"role": role}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Потребителят не е намерен")
    
    return {"message": "Ролята е обновена"}

@api_router.get("/auth/users")
async def get_all_users(current_user: User = Depends(get_current_user)):
    if current_user.role not in ["owner", "manager"]:
        raise HTTPException(status_code=403, detail="Нямате права за тази операция")
    
    if not current_user.company_id:
        return []
    
    users = await db.users.find(
        {"company_id": current_user.company_id},
        {"_id": 0, "user_id": 1, "email": 1, "name": 1, "role": 1, "picture": 1, "created_at": 1}
    ).to_list(1000)
    return users

@api_router.delete("/auth/users/{user_id}")
async def remove_user_from_company(user_id: str, current_user: User = Depends(get_current_user)):
    """Премахва потребител от фирмата (само Owner)"""
    if current_user.role != "owner":
        raise HTTPException(status_code=403, detail="Само титулярят може да премахва потребители")
    
    if user_id == current_user.user_id:
        raise HTTPException(status_code=400, detail="Не можете да премахнете себе си")
    
    target_user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="Потребителят не е намерен")
    
    if target_user.get("company_id") != current_user.company_id:
        raise HTTPException(status_code=403, detail="Потребителят не е от вашата фирма")
    
    # Remove company_id from user (don't delete user)
    await db.users.update_one(
        {"user_id": user_id},
        {"$unset": {"company_id": ""}, "$set": {"role": "staff"}}
    )
    
    return {"message": "Потребителят е премахнат от фирмата"}

# ===================== INVITATION ENDPOINTS =====================

@api_router.post("/invitations")
async def create_invitation(invitation_data: InvitationCreate, current_user: User = Depends(get_current_user)):
    """Създава покана за нов потребител (само Owner)"""
    if current_user.role != "owner":
        raise HTTPException(status_code=403, detail="Само титулярят може да изпраща покани")
    
    if not current_user.company_id:
        raise HTTPException(status_code=400, detail="Нямате фирма")
    
    if not invitation_data.email and not invitation_data.phone:
        raise HTTPException(status_code=400, detail="Въведете имейл или телефон")
    
    if invitation_data.role not in ["manager", "staff"]:
        raise HTTPException(status_code=400, detail="Невалидна роля за покана")
    
    # Check if user with this email already exists in the company
    if invitation_data.email:
        existing_user = await db.users.find_one({
            "email": invitation_data.email,
            "company_id": current_user.company_id
        })
        if existing_user:
            raise HTTPException(status_code=400, detail="Потребител с този имейл вече е член на фирмата")
    
    # Check for pending invitation
    pending = await db.invitations.find_one({
        "company_id": current_user.company_id,
        "$or": [
            {"email": invitation_data.email} if invitation_data.email else {"email": None},
            {"phone": invitation_data.phone} if invitation_data.phone else {"phone": None}
        ],
        "status": "pending"
    })
    if pending:
        raise HTTPException(status_code=400, detail="Вече има активна покана за този контакт")
    
    invitation = Invitation(
        company_id=current_user.company_id,
        invited_by=current_user.user_id,
        email=invitation_data.email,
        phone=invitation_data.phone,
        role=invitation_data.role
    )
    
    await db.invitations.insert_one(invitation.dict())
    
    # Get company name for response
    company = await db.companies.find_one({"id": current_user.company_id}, {"name": 1})
    company_name = company.get("name", "Unknown") if company else "Unknown"
    
    return {
        "message": "Поканата е създадена",
        "invitation": {
            "id": invitation.id,
            "code": invitation.code,
            "expires_at": invitation.expires_at.isoformat(),
            "company_name": company_name
        }
    }

@api_router.get("/invitations")
async def get_invitations(current_user: User = Depends(get_current_user)):
    """Връща всички покани за фирмата (само Owner)"""
    if current_user.role != "owner":
        raise HTTPException(status_code=403, detail="Само титулярят може да вижда покани")
    
    if not current_user.company_id:
        return []
    
    invitations = await db.invitations.find(
        {"company_id": current_user.company_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return invitations

@api_router.delete("/invitations/{invitation_id}")
async def cancel_invitation(invitation_id: str, current_user: User = Depends(get_current_user)):
    """Отменя покана (само Owner)"""
    if current_user.role != "owner":
        raise HTTPException(status_code=403, detail="Само титулярят може да отменя покани")
    
    invitation = await db.invitations.find_one({
        "id": invitation_id,
        "company_id": current_user.company_id
    })
    
    if not invitation:
        raise HTTPException(status_code=404, detail="Поканата не е намерена")
    
    if invitation["status"] != "pending":
        raise HTTPException(status_code=400, detail="Поканата вече не е активна")
    
    await db.invitations.update_one(
        {"id": invitation_id},
        {"$set": {"status": "cancelled"}}
    )
    
    return {"message": "Поканата е отменена"}

@api_router.post("/invitations/accept")
async def accept_invitation(request: Request, current_user: User = Depends(get_current_user)):
    """Приема покана по код"""
    body = await request.json()
    code = body.get("code", "").upper().strip()
    
    if not code:
        raise HTTPException(status_code=400, detail="Въведете код на поканата")
    
    # Find valid invitation
    invitation = await db.invitations.find_one({
        "code": code,
        "status": "pending"
    })
    
    if not invitation:
        raise HTTPException(status_code=404, detail="Невалиден или изтекъл код")
    
    # Check expiry
    expires_at = invitation["expires_at"]
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    
    if expires_at < datetime.now(timezone.utc):
        await db.invitations.update_one(
            {"id": invitation["id"]},
            {"$set": {"status": "expired"}}
        )
        raise HTTPException(status_code=400, detail="Поканата е изтекла")
    
    # Check if user already has a company
    if current_user.company_id:
        raise HTTPException(status_code=400, detail="Вече сте член на фирма. Първо напуснете текущата фирма.")
    
    # Accept invitation - link user to company
    await db.users.update_one(
        {"user_id": current_user.user_id},
        {"$set": {
            "company_id": invitation["company_id"],
            "role": invitation["role"]
        }}
    )
    
    # Mark invitation as accepted
    await db.invitations.update_one(
        {"id": invitation["id"]},
        {"$set": {"status": "accepted"}}
    )
    
    # Get company info
    company = await db.companies.find_one({"id": invitation["company_id"]}, {"_id": 0})
    
    return {
        "message": f"Успешно се присъединихте към {company['name'] if company else 'фирмата'}",
        "company": company
    }

@api_router.post("/company/leave")
async def leave_company(current_user: User = Depends(get_current_user)):
    """Напускане на фирма (не може Owner)"""
    if not current_user.company_id:
        raise HTTPException(status_code=400, detail="Не сте член на фирма")
    
    if current_user.role == "owner":
        raise HTTPException(status_code=400, detail="Титулярят не може да напусне фирмата. Прехвърлете собствеността първо.")
    
    await db.users.update_one(
        {"user_id": current_user.user_id},
        {"$unset": {"company_id": ""}, "$set": {"role": "staff"}}
    )
    
    return {"message": "Успешно напуснахте фирмата"}

# ===================== NOTIFICATION SETTINGS ENDPOINTS =====================

@api_router.get("/notifications/settings", response_model=NotificationSettings)
async def get_notification_settings(current_user: User = Depends(get_current_user)):
    settings = await db.notification_settings.find_one({"user_id": current_user.user_id}, {"_id": 0})
    if not settings:
        # Create default settings
        default_settings = NotificationSettings(user_id=current_user.user_id)
        await db.notification_settings.insert_one(default_settings.dict())
        return default_settings
    return NotificationSettings(**settings)

@api_router.put("/notifications/settings", response_model=NotificationSettings)
async def update_notification_settings(
    settings_update: NotificationSettingsUpdate,
    current_user: User = Depends(get_current_user)
):
    update_data = {k: v for k, v in settings_update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    existing = await db.notification_settings.find_one({"user_id": current_user.user_id}, {"_id": 0})
    
    if not existing:
        # Create new settings
        new_settings = NotificationSettings(user_id=current_user.user_id, **update_data)
        await db.notification_settings.insert_one(new_settings.dict())
        return new_settings
    
    await db.notification_settings.update_one(
        {"user_id": current_user.user_id},
        {"$set": update_data}
    )
    
    updated = await db.notification_settings.find_one({"user_id": current_user.user_id}, {"_id": 0})
    return NotificationSettings(**updated)

# ===================== COMPANY ENDPOINTS =====================

@api_router.post("/company", response_model=Company)
async def create_or_update_company(company_data: CompanyCreate, current_user: User = Depends(get_current_user)):
    """Създава нова фирма или обновява съществуваща (само Owner може да редактира)"""
    
    # Check if user already has a company
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0})
    
    if user_doc and user_doc.get("company_id"):
        # User has a company - only owner can edit
        if user_doc.get("role") != "owner":
            raise HTTPException(status_code=403, detail="Само титулярят може да редактира данните на фирмата")
        
        # Update existing company
        update_data = {k: v for k, v in company_data.dict().items() if v is not None}
        update_data["updated_at"] = datetime.now(timezone.utc)
        
        # Don't allow EIK change if company has other users
        other_users = await db.users.count_documents({
            "company_id": user_doc["company_id"],
            "user_id": {"$ne": current_user.user_id}
        })
        if other_users > 0 and "eik" in update_data:
            existing = await db.companies.find_one({"id": user_doc["company_id"]})
            if existing and existing.get("eik") != update_data["eik"]:
                raise HTTPException(status_code=400, detail="Не може да се промени ЕИК на фирма с други потребители")
        
        await db.companies.update_one(
            {"id": user_doc["company_id"]},
            {"$set": update_data}
        )
        
        updated_company = await db.companies.find_one({"id": user_doc["company_id"]}, {"_id": 0})
        return Company(**updated_company)
    else:
        # Check if company with this EIK already exists
        existing_company = await db.companies.find_one({"eik": company_data.eik}, {"_id": 0})
        
        if existing_company:
            raise HTTPException(status_code=400, detail="Фирма с този ЕИК вече съществува. Използвайте код за присъединяване.")
        
        # Create new company
        company = Company(**company_data.dict())
        await db.companies.insert_one(company.dict())
        
        # Link current user to this company as owner
        await db.users.update_one(
            {"user_id": current_user.user_id},
            {"$set": {"company_id": company.id, "role": "owner"}}
        )
        
        return company

@api_router.get("/company", response_model=Optional[Company])
async def get_my_company(current_user: User = Depends(get_current_user)):
    """Връща фирмата на текущия потребител"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0})
    
    if not user_doc or not user_doc.get("company_id"):
        return None
    
    company = await db.companies.find_one({"id": user_doc["company_id"]}, {"_id": 0})
    if not company:
        return None
    
    return Company(**company)

@api_router.put("/company", response_model=Company)
async def update_company(company_update: CompanyUpdate, current_user: User = Depends(get_current_user)):
    """Обновява фирмата на текущия потребител"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0})
    
    if not user_doc or not user_doc.get("company_id"):
        raise HTTPException(status_code=404, detail="Нямате свързана фирма. Първо създайте фирма.")
    
    update_data = {k: v for k, v in company_update.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Няма данни за обновяване")
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.companies.update_one(
        {"id": user_doc["company_id"]},
        {"$set": update_data}
    )
    
    updated_company = await db.companies.find_one({"id": user_doc["company_id"]}, {"_id": 0})
    return Company(**updated_company)

@api_router.post("/company/join/{eik}")
async def join_company_by_eik(eik: str, current_user: User = Depends(get_current_user)):
    """Присъединява потребител към съществуваща фирма по ЕИК"""
    company = await db.companies.find_one({"eik": eik}, {"_id": 0})
    
    if not company:
        raise HTTPException(status_code=404, detail=f"Фирма с ЕИК {eik} не е намерена")
    
    await db.users.update_one(
        {"user_id": current_user.user_id},
        {"$set": {"company_id": company["id"]}}
    )
    
    return {"message": f"Успешно се присъединихте към {company['name']}", "company": Company(**company)}

@api_router.get("/company/users")
async def get_company_users(current_user: User = Depends(get_current_user)):
    """Връща всички потребители от същата фирма"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0})
    
    if not user_doc or not user_doc.get("company_id"):
        return []
    
    users = await db.users.find(
        {"company_id": user_doc["company_id"]},
        {"_id": 0, "user_id": 1, "email": 1, "name": 1, "role": 1, "picture": 1}
    ).to_list(1000)
    
    return users

# ===================== AI DATA CORRECTION MODULE =====================

class DataCorrectionResult(BaseModel):
    original: dict
    corrected: dict
    corrections_made: List[str] = []
    confidence: float = 1.0

async def normalize_supplier_name(supplier: str, company_id: Optional[str] = None) -> tuple[str, bool]:
    """Нормализира име на доставчик и го съпоставя с известни доставчици"""
    if not supplier:
        return supplier, False
    
    # Почистване на основни проблеми
    supplier = supplier.strip()
    
    # Премахване на типични OCR грешки
    ocr_fixes = {
        '0': 'О',  # Нула -> О (за български текст)
        '1': 'І',  # Единица -> И (в някои контексти)
        '|': 'І',
        '!': 'І',
    }
    
    # Нормализиране на правните форми
    legal_forms = [
        (r'\bЕООД\b', 'ЕООД'),
        (r'\bООД\b', 'ООД'),
        (r'\bАД\b', 'АД'),
        (r'\bЕАД\b', 'ЕАД'),
        (r'\bЕТ\b', 'ЕТ'),
        (r'\bСД\b', 'СД'),
        (r'\bКД\b', 'КД'),
        (r'\bКДА\b', 'КДА'),
    ]
    
    normalized = supplier.upper()
    for pattern, replacement in legal_forms:
        normalized = re.sub(pattern, replacement, normalized, flags=re.IGNORECASE)
    
    # Ако има company_id, търсим съществуващ подобен доставчик
    if company_id:
        existing_suppliers = await db.invoices.distinct("supplier", {"company_id": company_id})
        
        # Fuzzy matching - търсим най-близкото съвпадение
        best_match = None
        best_score = 0
        
        for existing in existing_suppliers:
            # Проста метрика за сходство
            existing_norm = existing.upper().replace(' ', '').replace('.', '')
            supplier_norm = normalized.replace(' ', '').replace('.', '')
            
            # Проверка за частично съвпадение
            if existing_norm in supplier_norm or supplier_norm in existing_norm:
                score = len(min(existing_norm, supplier_norm, key=len)) / len(max(existing_norm, supplier_norm, key=len))
                if score > best_score and score > 0.7:
                    best_score = score
                    best_match = existing
            
            # Точно съвпадение с игнориране на интервали и точки
            if existing_norm == supplier_norm:
                return existing, True
        
        if best_match and best_score > 0.8:
            return best_match, True
    
    # Форматиране - първа буква главна
    words = normalized.split()
    formatted_words = []
    for word in words:
        if word in ['ЕООД', 'ООД', 'АД', 'ЕАД', 'ЕТ', 'СД', 'КД', 'КДА', 'LIDL', 'BILLA', 'KAUFLAND']:
            formatted_words.append(word)
        else:
            formatted_words.append(word.capitalize())
    
    return ' '.join(formatted_words), False

def fix_ocr_number_errors(value: str) -> str:
    """Поправя типични OCR грешки в числа"""
    if not value:
        return value
    
    # Типични OCR грешки при числа
    fixes = {
        'O': '0',
        'o': '0',
        'О': '0',  # Кирилица О
        'о': '0',  # Кирилица о
        'l': '1',
        'I': '1',
        'І': '1',  # Кирилица І
        '|': '1',
        'S': '5',
        's': '5',
        'B': '8',
        'Z': '2',
        'z': '2',
        ',': '.',  # Запетая към точка за десетични
    }
    
    result = value
    for wrong, correct in fixes.items():
        result = result.replace(wrong, correct)
    
    # Премахване на всичко освен цифри и точка
    result = re.sub(r'[^\d.]', '', result)
    
    return result

def parse_amount(value) -> float:
    """Парсва сума от различни формати"""
    if value is None:
        return 0.0
    
    if isinstance(value, (int, float)):
        return float(value)
    
    if isinstance(value, str):
        # Почистване
        cleaned = fix_ocr_number_errors(value)
        
        # Обработка на европейски формат (1.234,56 -> 1234.56)
        if ',' in value and '.' in value:
            if value.rfind(',') > value.rfind('.'):
                # Европейски формат: 1.234,56
                cleaned = value.replace('.', '').replace(',', '.')
        elif ',' in value:
            # Може да е десетична запетая
            cleaned = value.replace(',', '.')
        
        cleaned = re.sub(r'[^\d.]', '', cleaned)
        
        try:
            return float(cleaned) if cleaned else 0.0
        except ValueError:
            return 0.0
    
    return 0.0

def normalize_invoice_number(invoice_number: str) -> str:
    """Нормализира номер на фактура"""
    if not invoice_number:
        return invoice_number
    
    # Почистване от OCR грешки
    cleaned = invoice_number.strip()
    
    # Поправка на типични грешки
    ocr_fixes = {
        'O': '0',
        'o': '0',
        'l': '1',
        'I': '1',
    }
    
    # За номера на фактури, само в числовите части
    parts = re.split(r'(\D+)', cleaned)
    result_parts = []
    
    for part in parts:
        if part.isdigit() or re.match(r'^[\dOoIl]+$', part):
            # Числова част - поправяме OCR грешки
            fixed = part
            for wrong, correct in ocr_fixes.items():
                fixed = fixed.replace(wrong, correct)
            result_parts.append(fixed)
        else:
            result_parts.append(part.upper())
    
    return ''.join(result_parts)

def normalize_date(date_str: str) -> Optional[str]:
    """Нормализира дата в ISO формат YYYY-MM-DD"""
    if not date_str:
        return None
    
    # Почистване
    date_str = date_str.strip()
    
    # Поправка на OCR грешки в числата
    date_str = fix_ocr_number_errors(date_str)
    
    # Различни формати
    patterns = [
        (r'(\d{4})-(\d{1,2})-(\d{1,2})', '%Y-%m-%d'),  # 2024-01-15
        (r'(\d{1,2})\.(\d{1,2})\.(\d{4})', '%d.%m.%Y'),  # 15.01.2024
        (r'(\d{1,2})/(\d{1,2})/(\d{4})', '%d/%m/%Y'),  # 15/01/2024
        (r'(\d{1,2})-(\d{1,2})-(\d{4})', '%d-%m-%Y'),  # 15-01-2024
    ]
    
    for pattern, fmt in patterns:
        match = re.search(pattern, date_str)
        if match:
            try:
                parsed = datetime.strptime(match.group(), fmt)
                return parsed.strftime('%Y-%m-%d')
            except ValueError:
                continue
    
    return None

async def correct_ocr_data(
    data: dict,
    company_id: Optional[str] = None
) -> DataCorrectionResult:
    """
    AI-powered корекция на OCR данни
    """
    original = data.copy()
    corrected = data.copy()
    corrections = []
    
    # 1. Корекция на доставчик
    if data.get("supplier"):
        normalized_supplier, was_matched = await normalize_supplier_name(
            data["supplier"], 
            company_id
        )
        if normalized_supplier != data["supplier"]:
            corrected["supplier"] = normalized_supplier
            if was_matched:
                corrections.append(f"Доставчик съпоставен: '{data['supplier']}' → '{normalized_supplier}'")
            else:
                corrections.append(f"Доставчик нормализиран: '{data['supplier']}' → '{normalized_supplier}'")
    
    # 2. Корекция на номер на фактура
    if data.get("invoice_number"):
        normalized_number = normalize_invoice_number(data["invoice_number"])
        if normalized_number != data["invoice_number"]:
            corrected["invoice_number"] = normalized_number
            corrections.append(f"Номер на фактура коригиран: '{data['invoice_number']}' → '{normalized_number}'")
    
    # 3. Корекция на дата
    if data.get("invoice_date"):
        normalized_date = normalize_date(str(data["invoice_date"]))
        if normalized_date and normalized_date != data.get("invoice_date"):
            corrected["invoice_date"] = normalized_date
            corrections.append(f"Дата нормализирана: '{data['invoice_date']}' → '{normalized_date}'")
    
    # 4. Корекция на суми
    amount_fields = ["amount_without_vat", "vat_amount", "total_amount"]
    for field in amount_fields:
        if field in data:
            parsed = parse_amount(data[field])
            if parsed != data.get(field):
                corrected[field] = parsed
                corrections.append(f"{field} коригирано: '{data[field]}' → {parsed}")
    
    # 5. Валидация на ДДС изчисления
    amount_without_vat = corrected.get("amount_without_vat", 0)
    vat_amount = corrected.get("vat_amount", 0)
    total_amount = corrected.get("total_amount", 0)
    
    # Проверка за консистентност
    if amount_without_vat > 0 and total_amount > 0:
        expected_vat = amount_without_vat * 0.20
        expected_total = amount_without_vat + expected_vat
        
        # Ако ДДС е грешно, коригираме
        if vat_amount == 0 and total_amount > amount_without_vat:
            corrected["vat_amount"] = round(total_amount - amount_without_vat, 2)
            corrections.append(f"ДДС изчислено: {corrected['vat_amount']}")
        
        # Ако общата сума е грешна, коригираме
        elif abs(total_amount - (amount_without_vat + vat_amount)) > 0.02:
            corrected["total_amount"] = round(amount_without_vat + vat_amount, 2)
            corrections.append(f"Обща сума коригирана: {corrected['total_amount']}")
        
        # Ако само ДДС липсва
        elif vat_amount == 0 and total_amount == amount_without_vat:
            corrected["vat_amount"] = round(amount_without_vat * 0.20, 2)
            corrected["total_amount"] = round(amount_without_vat * 1.20, 2)
            corrections.append(f"ДДС добавено (20%): {corrected['vat_amount']}")
    
    # Изчисляване на confidence
    confidence = 1.0 - (len(corrections) * 0.05)  # Намаляме увереността с всяка корекция
    confidence = max(0.5, confidence)  # Минимум 50%
    
    return DataCorrectionResult(
        original=original,
        corrected=corrected,
        corrections_made=corrections,
        confidence=confidence
    )

# ===================== OCR ENDPOINT =====================

@api_router.post("/ocr/scan", response_model=OCRResult)
async def scan_invoice(image_base64: str = None, request: Request = None, current_user: User = Depends(get_current_user)):
    body = await request.json()
    image_data = body.get("image_base64", "")
    
    if not image_data:
        raise HTTPException(status_code=400, detail="Липсва изображение")
    
    # Remove data URL prefix if present
    if "," in image_data:
        image_data = image_data.split(",")[1]
    
    # Get company_id for supplier matching
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
        
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"ocr_{uuid.uuid4().hex[:8]}",
            system_message="""Ти си OCR асистент за извличане на данни от фактури на български език.
            Анализирай изображението и извлечи следните данни:
            - Доставчик (име на фирмата)
            - Номер на фактура
            - Дата на издаване на фактурата (във формат YYYY-MM-DD)
            - Сума без ДДС
            - ДДС (обикновено 20%)
            - Обща сума
            
            Отговори САМО в JSON формат:
            {"supplier": "...", "invoice_number": "...", "invoice_date": "YYYY-MM-DD", "amount_without_vat": 0.00, "vat_amount": 0.00, "total_amount": 0.00}
            
            Ако не можеш да прочетеш някоя стойност, използвай празен низ за текст или 0 за числа.
            За датата: ако не може да се прочете, върни null."""
        ).with_model("gemini", "gemini-2.5-flash")
        
        image_content = ImageContent(image_base64=image_data)
        
        user_message = UserMessage(
            text="Извлечи данните от тази фактура. Отговори само с JSON.",
            file_contents=[image_content]
        )
        
        response = await chat.send_message(user_message)
        
        # Parse JSON from response
        import json
        
        # Find JSON in response
        json_match = re.search(r'\{[^{}]*\}', response, re.DOTALL)
        if json_match:
            raw_result = json.loads(json_match.group())
            
            # Apply AI correction module
            correction_result = await correct_ocr_data(raw_result, company_id)
            corrected = correction_result.corrected
            
            # Log corrections for debugging
            if correction_result.corrections_made:
                logger.info(f"OCR Corrections: {correction_result.corrections_made}")
            
            return OCRResult(
                supplier=corrected.get("supplier", ""),
                invoice_number=corrected.get("invoice_number", ""),
                amount_without_vat=float(corrected.get("amount_without_vat", 0)),
                vat_amount=float(corrected.get("vat_amount", 0)),
                total_amount=float(corrected.get("total_amount", 0)),
                invoice_date=corrected.get("invoice_date"),
                corrections=correction_result.corrections_made,
                confidence=correction_result.confidence
            )
        else:
            raise HTTPException(status_code=500, detail="Не можах да разпозная фактурата")
            
    except Exception as e:
        logger.error(f"OCR Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Грешка при сканиране: {str(e)}")

# ===================== INVOICE ENDPOINTS =====================

@api_router.post("/invoices", response_model=Invoice)
async def create_invoice(invoice: InvoiceCreate, current_user: User = Depends(get_current_user)):
    # Get user's company_id
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    # Check for duplicate invoice
    if company_id:
        # If user has a company, check across all company users
        company_users = await db.users.find({"company_id": company_id}, {"user_id": 1}).to_list(1000)
        company_user_ids = [u["user_id"] for u in company_users]
        
        existing_invoice = await db.invoices.find_one({
            "user_id": {"$in": company_user_ids},
            "invoice_number": invoice.invoice_number,
            "supplier": {"$regex": f"^{invoice.supplier}$", "$options": "i"}
        }, {"_id": 0, "id": 1, "date": 1, "user_id": 1})
        
        if existing_invoice:
            # Find who added the duplicate
            added_by_user = await db.users.find_one({"user_id": existing_invoice["user_id"]}, {"name": 1})
            added_by_name = added_by_user.get("name", "друг потребител") if added_by_user else "друг потребител"
            raise HTTPException(
                status_code=409,
                detail=f"Фактура с номер {invoice.invoice_number} от {invoice.supplier} вече е добавена от {added_by_name}!"
            )
    else:
        # No company - check only for current user
        existing_invoice = await db.invoices.find_one({
            "user_id": current_user.user_id,
            "invoice_number": invoice.invoice_number,
            "supplier": {"$regex": f"^{invoice.supplier}$", "$options": "i"}
        }, {"_id": 0, "id": 1, "date": 1})
        
        if existing_invoice:
            raise HTTPException(
                status_code=409,
                detail=f"Фактура с номер {invoice.invoice_number} от {invoice.supplier} вече съществува в системата!"
            )
    
    invoice_dict = invoice.dict()
    invoice_date = datetime.fromisoformat(invoice_dict["date"].replace("Z", "+00:00"))
    
    # Process items and convert to dict format
    items_list = None
    price_alerts = []
    
    if invoice.items:
        items_list = []
        for item in invoice.items:
            item_dict = item.dict()
            # Calculate total_price if not provided
            if item_dict.get("total_price") is None:
                item_dict["total_price"] = item_dict["quantity"] * item_dict["unit_price"]
            # Calculate VAT if not provided (20%)
            if item_dict.get("vat_amount") is None:
                item_dict["vat_amount"] = item_dict["total_price"] * 0.2
            
            item_dict["id"] = str(uuid.uuid4())
            items_list.append(item_dict)
            
            # Check price changes and create alerts if company exists
            if company_id:
                normalized_name = item.name.strip().lower()
                
                # Find last price for this item from same supplier
                last_price_record = await db.item_price_history.find_one(
                    {
                        "company_id": company_id,
                        "supplier": {"$regex": f"^{invoice.supplier}$", "$options": "i"},
                        "item_name": normalized_name
                    },
                    {"_id": 0},
                    sort=[("invoice_date", -1)]
                )
                
                # Get threshold setting
                alert_settings = await db.price_alert_settings.find_one(
                    {"company_id": company_id},
                    {"_id": 0}
                )
                threshold = alert_settings.get("threshold_percent", 10.0) if alert_settings else 10.0
                alert_enabled = alert_settings.get("enabled", True) if alert_settings else True
                
                if last_price_record and alert_enabled:
                    old_price = last_price_record["unit_price"]
                    new_price = item.unit_price
                    
                    if old_price > 0:
                        change_percent = ((new_price - old_price) / old_price) * 100
                        
                        # Create alert if price increased above threshold
                        if change_percent >= threshold:
                            alert = PriceAlert(
                                company_id=company_id,
                                item_name=item.name,
                                supplier=invoice.supplier,
                                old_price=old_price,
                                new_price=new_price,
                                change_percent=round(change_percent, 2),
                                invoice_id="",  # Will be set after invoice is created
                                invoice_number=invoice.invoice_number
                            )
                            price_alerts.append(alert)
                
                # Save to price history
                price_history = ItemPriceHistory(
                    company_id=company_id,
                    supplier=invoice.supplier,
                    item_name=normalized_name,
                    unit_price=item.unit_price,
                    quantity=item.quantity,
                    unit=item.unit,
                    invoice_id="",  # Will be set after invoice is created
                    invoice_number=invoice.invoice_number,
                    invoice_date=invoice_date
                )
                await db.item_price_history.insert_one(price_history.dict())
    
    invoice_obj = Invoice(
        user_id=current_user.user_id,
        company_id=company_id,
        date=invoice_date,
        items=items_list,
        **{k: v for k, v in invoice_dict.items() if k not in ["date", "items"]}
    )
    await db.invoices.insert_one(invoice_obj.dict())
    
    # Update price history and alerts with invoice_id
    if company_id and invoice.items:
        for item in invoice.items:
            normalized_name = item.name.strip().lower()
            await db.item_price_history.update_many(
                {
                    "company_id": company_id,
                    "invoice_number": invoice.invoice_number,
                    "item_name": normalized_name,
                    "invoice_id": ""
                },
                {"$set": {"invoice_id": invoice_obj.id}}
            )
        
        # Save alerts with invoice_id
        for alert in price_alerts:
            alert.invoice_id = invoice_obj.id
            await db.price_alerts.insert_one(alert.dict())
    
    return invoice_obj

@api_router.get("/invoices", response_model=List[Invoice])
async def get_invoices(
    supplier: Optional[str] = None,
    invoice_number: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {"user_id": current_user.user_id}
    
    if supplier:
        query["supplier"] = {"$regex": supplier, "$options": "i"}
    if invoice_number:
        query["invoice_number"] = {"$regex": invoice_number, "$options": "i"}
    if start_date:
        query["date"] = {"$gte": datetime.fromisoformat(start_date.replace("Z", "+00:00"))}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
        else:
            query["date"] = {"$lte": datetime.fromisoformat(end_date.replace("Z", "+00:00"))}
    
    invoices = await db.invoices.find(query, {"_id": 0, "image_base64": 0}).sort("date", -1).to_list(1000)
    return [Invoice(**inv) for inv in invoices]

@api_router.get("/invoices/{invoice_id}", response_model=Invoice)
async def get_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    invoice = await db.invoices.find_one({"id": invoice_id, "user_id": current_user.user_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Фактурата не е намерена")
    return Invoice(**invoice)

@api_router.put("/invoices/{invoice_id}", response_model=Invoice)
async def update_invoice(invoice_id: str, invoice_update: InvoiceUpdate, current_user: User = Depends(get_current_user)):
    update_data = {k: v for k, v in invoice_update.dict().items() if v is not None}
    if "date" in update_data:
        update_data["date"] = datetime.fromisoformat(update_data["date"].replace("Z", "+00:00"))
    
    result = await db.invoices.update_one(
        {"id": invoice_id, "user_id": current_user.user_id},
        {"$set": update_data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Фактурата не е намерена")
    
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    return Invoice(**invoice)

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    result = await db.invoices.delete_one({"id": invoice_id, "user_id": current_user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Фактурата не е намерена")
    return {"message": "Фактурата е изтрита"}

# ===================== DAILY REVENUE ENDPOINTS =====================

@api_router.post("/daily-revenue", response_model=DailyRevenue)
async def create_daily_revenue(revenue: DailyRevenueCreate, current_user: User = Depends(get_current_user)):
    # Check if entry for this date already exists
    existing = await db.daily_revenue.find_one({
        "user_id": current_user.user_id,
        "date": revenue.date
    }, {"_id": 0})
    
    if existing:
        # ADD to existing values instead of replacing
        new_fiscal = existing.get("fiscal_revenue", 0) + revenue.fiscal_revenue
        new_pocket = existing.get("pocket_money", 0) + revenue.pocket_money
        
        await db.daily_revenue.update_one(
            {"id": existing["id"]},
            {"$set": {"fiscal_revenue": new_fiscal, "pocket_money": new_pocket}}
        )
        existing["fiscal_revenue"] = new_fiscal
        existing["pocket_money"] = new_pocket
        return DailyRevenue(**existing)
    
    revenue_obj = DailyRevenue(
        user_id=current_user.user_id,
        date=revenue.date,
        fiscal_revenue=revenue.fiscal_revenue,
        pocket_money=revenue.pocket_money
    )
    await db.daily_revenue.insert_one(revenue_obj.dict())
    return revenue_obj

@api_router.get("/daily-revenue/today")
async def get_today_revenue(current_user: User = Depends(get_current_user)):
    """Get today's revenue totals"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    existing = await db.daily_revenue.find_one({
        "user_id": current_user.user_id,
        "date": today
    }, {"_id": 0})
    
    if existing:
        return {
            "date": today,
            "fiscal_revenue": existing.get("fiscal_revenue", 0),
            "pocket_money": existing.get("pocket_money", 0)
        }
    return {
        "date": today,
        "fiscal_revenue": 0,
        "pocket_money": 0
    }

@api_router.get("/daily-revenue/by-date/{date}")
async def get_revenue_by_date(date: str, current_user: User = Depends(get_current_user)):
    """Get revenue for a specific date"""
    existing = await db.daily_revenue.find_one({
        "user_id": current_user.user_id,
        "date": date
    }, {"_id": 0})
    
    if existing:
        return {
            "date": date,
            "fiscal_revenue": existing.get("fiscal_revenue", 0),
            "pocket_money": existing.get("pocket_money", 0)
        }
    return {
        "date": date,
        "fiscal_revenue": 0,
        "pocket_money": 0
    }

@api_router.get("/daily-revenue", response_model=List[DailyRevenue])
async def get_daily_revenues(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {"user_id": current_user.user_id}
    
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    revenues = await db.daily_revenue.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return [DailyRevenue(**r) for r in revenues]

# ===================== NON-INVOICE EXPENSE ENDPOINTS =====================

@api_router.post("/expenses", response_model=NonInvoiceExpense)
async def create_expense(expense: NonInvoiceExpenseCreate, current_user: User = Depends(get_current_user)):
    expense_obj = NonInvoiceExpense(
        user_id=current_user.user_id,
        **expense.dict()
    )
    await db.expenses.insert_one(expense_obj.dict())
    return expense_obj

@api_router.get("/expenses", response_model=List[NonInvoiceExpense])
async def get_expenses(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {"user_id": current_user.user_id}
    
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    expenses = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return [NonInvoiceExpense(**e) for e in expenses]

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, current_user: User = Depends(get_current_user)):
    result = await db.expenses.delete_one({"id": expense_id, "user_id": current_user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Разходът не е намерен")
    return {"message": "Разходът е изтрит"}

# ===================== PERSONAL EXPENSES & ROI ENDPOINTS =====================

@api_router.post("/personal-expenses")
async def create_personal_expense(
    expense: PersonalExpenseCreate,
    current_user: User = Depends(get_current_user)
):
    """Създава личен разход/инвестиция (само за собственик)"""
    # Check if user is owner
    if current_user.role != "owner":
        raise HTTPException(status_code=403, detail="Само титулярът може да управлява лични разходи")
    
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        raise HTTPException(status_code=400, detail="Няма свързана фирма")
    
    expense_obj = PersonalExpense(
        user_id=current_user.user_id,
        company_id=company_id,
        amount=expense.amount,
        description=expense.description,
        expense_type=PersonalExpenseType(expense.expense_type),
        category=PersonalExpenseCategory(expense.category),
        period_month=expense.period_month,
        period_year=expense.period_year,
        supplier_id=expense.supplier_id,
        project_name=expense.project_name,
        notes=expense.notes
    )
    
    await db.personal_expenses.insert_one(expense_obj.dict())
    return {"message": "Личният разход е записан", "id": expense_obj.id}

@api_router.get("/personal-expenses")
async def get_personal_expenses(
    month: Optional[int] = None,
    year: Optional[int] = None,
    expense_type: Optional[str] = None,
    category: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Връща личните разходи (само за собственик)"""
    if current_user.role != "owner":
        raise HTTPException(status_code=403, detail="Само титулярът може да вижда лични разходи")
    
    query = {"user_id": current_user.user_id}
    
    if month:
        query["period_month"] = month
    if year:
        query["period_year"] = year
    if expense_type:
        query["expense_type"] = expense_type
    if category:
        query["category"] = category
    
    expenses = await db.personal_expenses.find(query, {"_id": 0}).sort([("period_year", -1), ("period_month", -1)]).to_list(1000)
    return {"personal_expenses": expenses}

@api_router.delete("/personal-expenses/{expense_id}")
async def delete_personal_expense(
    expense_id: str,
    current_user: User = Depends(get_current_user)
):
    """Изтрива личен разход (само за собственик)"""
    if current_user.role != "owner":
        raise HTTPException(status_code=403, detail="Само титулярът може да управлява лични разходи")
    
    result = await db.personal_expenses.delete_one({
        "id": expense_id,
        "user_id": current_user.user_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Разходът не е намерен")
    
    return {"message": "Личният разход е изтрит"}

@api_router.get("/roi/analysis")
async def get_roi_analysis(
    month: Optional[int] = None,
    year: Optional[int] = None,
    current_user: User = Depends(get_current_user)
):
    """
    ROI анализ - изчислява възвръщаемост на личната инвестиция.
    Само за собственик.
    """
    if current_user.role != "owner":
        raise HTTPException(status_code=403, detail="Само титулярът има достъп до ROI анализ")
    
    # Default to current month/year
    now = datetime.now(timezone.utc)
    target_month = month or now.month
    target_year = year or now.year
    
    # Build period dates
    period_start = f"{target_year}-{target_month:02d}-01"
    if target_month == 12:
        period_end = f"{target_year + 1}-01-01"
    else:
        period_end = f"{target_year}-{target_month + 1:02d}-01"
    
    # Get personal expenses for period
    personal_expenses = await db.personal_expenses.find({
        "user_id": current_user.user_id,
        "period_month": target_month,
        "period_year": target_year
    }, {"_id": 0, "amount": 1, "expense_type": 1}).to_list(1000)
    
    total_personal = sum(e.get("amount", 0) for e in personal_expenses)
    total_investment = sum(e.get("amount", 0) for e in personal_expenses if e.get("expense_type") == "investment")
    
    # Get business revenue for period
    revenues = await db.daily_revenue.find({
        "user_id": current_user.user_id,
        "date": {"$gte": period_start, "$lt": period_end}
    }, {"_id": 0, "fiscal_revenue": 1, "pocket_money": 1}).to_list(1000)
    
    total_revenue = sum(r.get("fiscal_revenue", 0) + r.get("pocket_money", 0) for r in revenues)
    
    # Get business expenses (invoices + non-invoice expenses)
    invoices = await db.invoices.find({
        "user_id": current_user.user_id,
        "date": {
            "$gte": datetime.fromisoformat(period_start + "T00:00:00+00:00"),
            "$lt": datetime.fromisoformat(period_end + "T00:00:00+00:00")
        }
    }, {"_id": 0, "total_amount": 1}).to_list(1000)
    
    business_expenses = await db.expenses.find({
        "user_id": current_user.user_id,
        "date": {"$gte": period_start, "$lt": period_end}
    }, {"_id": 0, "amount": 1}).to_list(1000)
    
    total_business_expense = sum(i.get("total_amount", 0) for i in invoices) + sum(e.get("amount", 0) for e in business_expenses)
    
    # Calculate profit and ROI
    total_profit = total_revenue - total_business_expense
    
    # ROI = (Печалба / Лична инвестиция) * 100
    if total_personal > 0:
        roi_percent = (total_profit / total_personal) * 100
    else:
        roi_percent = 0 if total_profit <= 0 else 100
    
    is_profitable = total_profit > 0
    investment_covered = total_profit >= total_personal
    
    # Generate AI insights
    ai_insights = await generate_roi_insights(
        total_personal=total_personal,
        total_investment=total_investment,
        total_revenue=total_revenue,
        total_profit=total_profit,
        roi_percent=roi_percent,
        is_profitable=is_profitable,
        investment_covered=investment_covered
    )
    
    return {
        "period": {"month": target_month, "year": target_year},
        "period_start": period_start,
        "period_end": period_end,
        "total_personal_investment": round(total_personal, 2),
        "total_investment_only": round(total_investment, 2),
        "total_revenue": round(total_revenue, 2),
        "total_business_expense": round(total_business_expense, 2),
        "total_profit": round(total_profit, 2),
        "roi_percent": round(roi_percent, 1),
        "is_profitable": is_profitable,
        "investment_covered": investment_covered,
        "ai_insights": ai_insights
    }

async def generate_roi_insights(
    total_personal: float,
    total_investment: float,
    total_revenue: float,
    total_profit: float,
    roi_percent: float,
    is_profitable: bool,
    investment_covered: bool
) -> List[str]:
    """Генерира AI управленски предложения за ROI"""
    insights = []
    
    # Basic insights без AI (винаги налични)
    if total_personal == 0:
        insights.append("📊 Няма въведени лични разходи за периода")
        return insights
    
    if investment_covered:
        insights.append("✅ Бизнесът покрива личната инвестиция за периода")
    elif is_profitable:
        diff = total_personal - total_profit
        insights.append(f"⚠️ Печалбата не покрива напълно личната инвестиция (остават {diff:.2f} лв)")
    else:
        insights.append("❌ Работиш повече за бизнеса, отколкото бизнесът за теб")
    
    if roi_percent > 100:
        insights.append(f"🚀 Отличен ROI: {roi_percent:.1f}% - инвестицията се изплаща многократно")
    elif roi_percent > 50:
        insights.append(f"📈 Добър ROI: {roi_percent:.1f}%")
    elif roi_percent > 0:
        insights.append(f"📉 Нисък ROI: {roi_percent:.1f}% - има място за подобрение")
    elif roi_percent < -50:
        insights.append(f"🔴 Отрицателен ROI: {roi_percent:.1f}% - необходим е анализ на разходите")
    
    # Personal contribution ratio
    if total_revenue > 0:
        personal_ratio = (total_personal / total_revenue) * 100
        if personal_ratio > 50:
            insights.append(f"⚠️ Личната контрибуция ({personal_ratio:.1f}%) е висока спрямо оборота")
        elif personal_ratio > 30:
            insights.append(f"📊 Личната контрибуция е {personal_ratio:.1f}% от оборота")
    
    # Try to get AI enhanced insights
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        llm = LlmChat(
            api_key=os.getenv("EMERGENT_LLM_KEY"),
            session_id=f"roi_{uuid.uuid4().hex[:8]}",
            system_message="Ти си финансов съветник за малък бизнес. Давай кратки, ясни и практични съвети на български."
        ).with_model("gemini", "gemini-2.5-flash")
        
        prompt = f"""Анализирай тези финансови показатели за малък бизнес:
- Лична инвестиция на собственика: {total_personal:.2f} лв
- Общ оборот: {total_revenue:.2f} лв
- Печалба: {total_profit:.2f} лв
- ROI: {roi_percent:.1f}%

Дай ЕДНА кратка препоръка (до 15 думи) какво може да направи собственикът за подобрение.
Отговори директно с препоръката, без въвеждащ текст."""

        response = await llm.send_message(UserMessage(text=prompt))
        ai_recommendation = response.strip() if isinstance(response, str) else str(response)
        
        if ai_recommendation and len(ai_recommendation) < 200:
            insights.append(f"💡 AI препоръка: {ai_recommendation}")
    except Exception as e:
        logger.warning(f"AI insights generation failed: {str(e)}")
    
    return insights

@api_router.get("/roi/trend")
async def get_roi_trend(
    months: int = 6,
    current_user: User = Depends(get_current_user)
):
    """Връща ROI тренд за последните N месеца (само за собственик)"""
    if current_user.role != "owner":
        raise HTTPException(status_code=403, detail="Само титулярът има достъп до ROI тренд")
    
    now = datetime.now(timezone.utc)
    trend_data = []
    
    for i in range(months - 1, -1, -1):
        # Calculate target month
        target_month = now.month - i
        target_year = now.year
        
        while target_month <= 0:
            target_month += 12
            target_year -= 1
        
        # Get ROI for this month
        period_start = f"{target_year}-{target_month:02d}-01"
        if target_month == 12:
            period_end = f"{target_year + 1}-01-01"
        else:
            period_end = f"{target_year}-{target_month + 1:02d}-01"
        
        # Personal expenses
        personal = await db.personal_expenses.find({
            "user_id": current_user.user_id,
            "period_month": target_month,
            "period_year": target_year
        }, {"_id": 0, "amount": 1}).to_list(1000)
        total_personal = sum(p.get("amount", 0) for p in personal)
        
        # Revenue
        revenues = await db.daily_revenue.find({
            "user_id": current_user.user_id,
            "date": {"$gte": period_start, "$lt": period_end}
        }, {"_id": 0, "fiscal_revenue": 1, "pocket_money": 1}).to_list(1000)
        total_revenue = sum(r.get("fiscal_revenue", 0) + r.get("pocket_money", 0) for r in revenues)
        
        # Business expenses
        invoices = await db.invoices.find({
            "user_id": current_user.user_id,
            "date": {
                "$gte": datetime.fromisoformat(period_start + "T00:00:00+00:00"),
                "$lt": datetime.fromisoformat(period_end + "T00:00:00+00:00")
            }
        }, {"_id": 0, "total_amount": 1}).to_list(1000)
        
        expenses = await db.expenses.find({
            "user_id": current_user.user_id,
            "date": {"$gte": period_start, "$lt": period_end}
        }, {"_id": 0, "amount": 1}).to_list(1000)
        
        total_expense = sum(i.get("total_amount", 0) for i in invoices) + sum(e.get("amount", 0) for e in expenses)
        total_profit = total_revenue - total_expense
        
        roi = (total_profit / total_personal * 100) if total_personal > 0 else (0 if total_profit <= 0 else 100)
        
        trend_data.append({
            "month": target_month,
            "year": target_year,
            "label": f"{target_month:02d}/{target_year}",
            "personal_investment": round(total_personal, 2),
            "revenue": round(total_revenue, 2),
            "profit": round(total_profit, 2),
            "roi_percent": round(roi, 1)
        })
    
    return {"trend": trend_data, "months": months}

# ===================== STATISTICS ENDPOINTS =====================

@api_router.get("/statistics/summary")
async def get_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_month_only: bool = True,  # Default to current month
    current_user: User = Depends(get_current_user)
):
    # If no dates provided and current_month_only is True, use current month
    if not start_date and not end_date and current_month_only:
        now = datetime.now(timezone.utc)
        start_date = now.replace(day=1).strftime("%Y-%m-%d")
        # End of month
        if now.month == 12:
            end_date = now.replace(year=now.year + 1, month=1, day=1).strftime("%Y-%m-%d")
        else:
            end_date = now.replace(month=now.month + 1, day=1).strftime("%Y-%m-%d")
    
    date_query = {}
    
    if start_date:
        date_query["$gte"] = start_date
    if end_date:
        date_query["$lte"] = end_date
    
    # Get invoices
    inv_query = {"user_id": current_user.user_id}
    if start_date or end_date:
        inv_query["date"] = {}
        if start_date:
            inv_query["date"]["$gte"] = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
        if end_date:
            inv_query["date"]["$lte"] = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
    
    invoices = await db.invoices.find(inv_query, {"_id": 0, "total_amount": 1, "vat_amount": 1}).to_list(1000)
    
    # Get daily revenues
    rev_query = {"user_id": current_user.user_id}
    if date_query:
        rev_query["date"] = date_query
    revenues = await db.daily_revenue.find(rev_query, {"_id": 0, "fiscal_revenue": 1, "pocket_money": 1}).to_list(1000)
    
    # Get expenses
    exp_query = {"user_id": current_user.user_id}
    if date_query:
        exp_query["date"] = date_query
    expenses = await db.expenses.find(exp_query, {"_id": 0, "amount": 1}).to_list(1000)
    
    # Calculate totals
    total_invoice_amount = sum(inv.get("total_amount", 0) for inv in invoices)
    total_invoice_vat = sum(inv.get("vat_amount", 0) for inv in invoices)
    total_fiscal_revenue = sum(r.get("fiscal_revenue", 0) for r in revenues)
    total_pocket_money = sum(r.get("pocket_money", 0) for r in revenues)
    total_expenses = sum(e.get("amount", 0) for e in expenses)
    
    # ДДС от фискализиран оборот (20% от 120% = 16.67% от тотала)
    fiscal_vat = total_fiscal_revenue * 0.2 / 1.2
    
    # Общ ДДС за плащане = ДДС от продажби - ДДС от покупки (фактури)
    vat_to_pay = fiscal_vat - total_invoice_vat
    
    # Общ приход (фискализиран + джобче)
    total_income = total_fiscal_revenue + total_pocket_money
    
    # Общ разход (фактури + разходи без фактури)
    total_expense = total_invoice_amount + total_expenses
    
    return {
        "total_invoice_amount": round(total_invoice_amount, 2),
        "total_invoice_vat": round(total_invoice_vat, 2),
        "total_fiscal_revenue": round(total_fiscal_revenue, 2),
        "total_pocket_money": round(total_pocket_money, 2),
        "fiscal_vat": round(fiscal_vat, 2),
        "vat_to_pay": round(vat_to_pay, 2),
        "total_non_invoice_expenses": round(total_expenses, 2),
        "total_income": round(total_income, 2),
        "total_expense": round(total_expense, 2),
        "profit": round(total_income - total_expense, 2),
        "invoice_count": len(invoices)
    }

@api_router.get("/statistics/chart-data")
async def get_chart_data(
    period: str = "week",  # week, month, year
    current_user: User = Depends(get_current_user)
):
    now = datetime.now(timezone.utc)
    
    if period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now - timedelta(days=30)
    else:
        start = now - timedelta(days=365)
    
    start_str = start.strftime("%Y-%m-%d")
    
    # Get data
    inv_query = {
        "user_id": current_user.user_id,
        "date": {"$gte": start}
    }
    invoices = await db.invoices.find(inv_query, {"_id": 0, "date": 1, "total_amount": 1, "vat_amount": 1}).to_list(1000)
    
    rev_query = {
        "user_id": current_user.user_id,
        "date": {"$gte": start_str}
    }
    revenues = await db.daily_revenue.find(rev_query, {"_id": 0, "date": 1, "fiscal_revenue": 1, "pocket_money": 1}).to_list(1000)
    
    exp_query = {
        "user_id": current_user.user_id,
        "date": {"$gte": start_str}
    }
    expenses = await db.expenses.find(exp_query, {"_id": 0, "date": 1, "amount": 1}).to_list(1000)
    
    # Group by date
    from collections import defaultdict
    
    daily_data = defaultdict(lambda: {"income": 0, "expense": 0, "vat": 0})
    
    for inv in invoices:
        date_str = inv["date"].strftime("%Y-%m-%d") if isinstance(inv["date"], datetime) else inv["date"][:10]
        daily_data[date_str]["expense"] += inv.get("total_amount", 0)
        daily_data[date_str]["vat"] -= inv.get("vat_amount", 0)  # ДДС кредит
    
    for rev in revenues:
        date_str = rev["date"][:10] if isinstance(rev["date"], str) else rev["date"].strftime("%Y-%m-%d")
        daily_data[date_str]["income"] += rev.get("fiscal_revenue", 0) + rev.get("pocket_money", 0)
        daily_data[date_str]["vat"] += rev.get("fiscal_revenue", 0) * 0.2 / 1.2  # ДДС от продажби
    
    for exp in expenses:
        date_str = exp["date"][:10] if isinstance(exp["date"], str) else exp["date"].strftime("%Y-%m-%d")
        daily_data[date_str]["expense"] += exp.get("amount", 0)
    
    # Convert to list sorted by date
    chart_data = []
    for date_str in sorted(daily_data.keys()):
        chart_data.append({
            "date": date_str,
            "label": date_str[5:],  # MM-DD
            "income": round(daily_data[date_str]["income"], 2),
            "expense": round(daily_data[date_str]["expense"], 2),
            "vat": round(daily_data[date_str]["vat"], 2)
        })
    
    return chart_data

@api_router.get("/statistics/suppliers")
async def get_supplier_statistics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get comprehensive supplier statistics with trends, alerts, and rankings"""
    from collections import defaultdict
    from datetime import timedelta
    
    # Default to current month if no dates provided
    now = datetime.now(timezone.utc)
    if not start_date and not end_date:
        start_date = now.replace(day=1).strftime("%Y-%m-%d")
        if now.month == 12:
            end_date = now.replace(year=now.year + 1, month=1, day=1).strftime("%Y-%m-%d")
        else:
            end_date = now.replace(month=now.month + 1, day=1).strftime("%Y-%m-%d")
    
    # Build query for current period
    query = {"user_id": current_user.user_id}
    if start_date or end_date:
        query["date"] = {}
        if start_date:
            query["date"]["$gte"] = datetime.fromisoformat(start_date + "T00:00:00+00:00")
        if end_date:
            query["date"]["$lte"] = datetime.fromisoformat(end_date + "T23:59:59+00:00")
    
    # Get all invoices for the period
    invoices = await db.invoices.find(query, {
        "_id": 0, 
        "supplier": 1, 
        "total_amount": 1, 
        "vat_amount": 1,
        "amount_without_vat": 1,
        "date": 1
    }).to_list(10000)
    
    # Group by supplier with detailed data
    supplier_data = defaultdict(lambda: {
        "total_amount": 0,
        "total_vat": 0,
        "total_net": 0,
        "invoice_count": 0,
        "dates": [],
        "amounts": []
    })
    
    for inv in invoices:
        supplier = inv.get("supplier", "Неизвестен")
        amount = inv.get("total_amount", 0)
        supplier_data[supplier]["total_amount"] += amount
        supplier_data[supplier]["total_vat"] += inv.get("vat_amount", 0)
        supplier_data[supplier]["total_net"] += inv.get("amount_without_vat", 0)
        supplier_data[supplier]["invoice_count"] += 1
        
        date_val = inv.get("date")
        if isinstance(date_val, datetime):
            supplier_data[supplier]["dates"].append(date_val)
            supplier_data[supplier]["amounts"].append(amount)
    
    # Calculate inactivity threshold (30 days)
    inactivity_days = 30
    inactive_threshold = now - timedelta(days=inactivity_days)
    
    # Build comprehensive supplier list
    suppliers_list = []
    total_all = 0
    
    for supplier, data in supplier_data.items():
        total_all += data["total_amount"]
        
        # Calculate first/last delivery
        sorted_dates = sorted(data["dates"]) if data["dates"] else []
        first_delivery = sorted_dates[0] if sorted_dates else None
        last_delivery = sorted_dates[-1] if sorted_dates else None
        
        # Make dates timezone-aware if they aren't
        if last_delivery and last_delivery.tzinfo is None:
            last_delivery = last_delivery.replace(tzinfo=timezone.utc)
        if first_delivery and first_delivery.tzinfo is None:
            first_delivery = first_delivery.replace(tzinfo=timezone.utc)
        
        # Determine status
        is_active = last_delivery and last_delivery > inactive_threshold if last_delivery else False
        days_inactive = (now - last_delivery).days if last_delivery else 999
        
        # Calculate average
        avg = data["total_amount"] / data["invoice_count"] if data["invoice_count"] > 0 else 0
        
        # Calculate standard deviation for anomaly detection
        amounts = data["amounts"]
        if len(amounts) >= 2:
            mean_amount = sum(amounts) / len(amounts)
            variance = sum((x - mean_amount) ** 2 for x in amounts) / len(amounts)
            std_dev = variance ** 0.5
        else:
            mean_amount = avg
            std_dev = 0
        
        suppliers_list.append({
            "supplier": supplier,
            "total_amount": round(data["total_amount"], 2),
            "total_vat": round(data["total_vat"], 2),
            "total_net": round(data["total_net"], 2),
            "invoice_count": data["invoice_count"],
            "avg_invoice": round(avg, 2),
            "first_delivery": first_delivery.strftime("%Y-%m-%d") if first_delivery else None,
            "last_delivery": last_delivery.strftime("%Y-%m-%d") if last_delivery else None,
            "is_active": is_active,
            "days_inactive": days_inactive if not is_active else 0,
            "std_dev": round(std_dev, 2)
        })
    
    # Add dependency percentage
    for s in suppliers_list:
        s["dependency_percent"] = round((s["total_amount"] / total_all * 100), 1) if total_all > 0 else 0
    
    # Sort by total amount for TOP by amount
    top_by_amount = sorted(suppliers_list, key=lambda x: x["total_amount"], reverse=True)[:10]
    
    # TOP by frequency
    top_by_frequency = sorted(suppliers_list, key=lambda x: x["invoice_count"], reverse=True)[:10]
    
    # TOP by average invoice
    top_by_avg = sorted(suppliers_list, key=lambda x: x["avg_invoice"], reverse=True)[:10]
    
    # Inactive suppliers
    inactive_suppliers = [s for s in suppliers_list if not s["is_active"]]
    inactive_suppliers.sort(key=lambda x: x["days_inactive"], reverse=True)
    
    # High dependency alert (>30%)
    high_dependency = [s for s in suppliers_list if s["dependency_percent"] > 30]
    
    # Calculate totals
    total_vat = sum(s["total_vat"] for s in suppliers_list)
    total_net = sum(s["total_net"] for s in suppliers_list)
    total_invoices = sum(s["invoice_count"] for s in suppliers_list)
    
    # Executive summary
    top_3_concentration = sum(s["dependency_percent"] for s in top_by_amount[:3]) if len(top_by_amount) >= 3 else 0
    top_5_concentration = sum(s["dependency_percent"] for s in top_by_amount[:5]) if len(top_by_amount) >= 5 else 0
    
    return {
        "period": {
            "start_date": start_date,
            "end_date": end_date
        },
        "executive_summary": {
            "top_3_concentration": round(top_3_concentration, 1),
            "top_5_concentration": round(top_5_concentration, 1),
            "total_suppliers": len(suppliers_list),
            "active_suppliers": len([s for s in suppliers_list if s["is_active"]]),
            "inactive_suppliers": len(inactive_suppliers),
            "high_dependency_count": len(high_dependency),
            "largest_supplier": top_by_amount[0]["supplier"] if top_by_amount else None,
            "largest_amount": top_by_amount[0]["total_amount"] if top_by_amount else 0
        },
        "totals": {
            "total_amount": round(total_all, 2),
            "total_vat": round(total_vat, 2),
            "total_net": round(total_net, 2),
            "supplier_count": len(suppliers_list),
            "invoice_count": total_invoices
        },
        "top_by_amount": top_by_amount,
        "top_by_frequency": top_by_frequency,
        "top_by_avg": top_by_avg,
        "inactive_suppliers": inactive_suppliers[:10],
        "high_dependency_alerts": high_dependency,
        "all_suppliers": suppliers_list
    }

@api_router.get("/statistics/supplier/{supplier_name}/detailed")
async def get_detailed_supplier_stats(
    supplier_name: str,
    current_user: User = Depends(get_current_user)
):
    """Get detailed statistics for a specific supplier with trends"""
    from collections import defaultdict
    from urllib.parse import unquote
    
    supplier_name = unquote(supplier_name)
    
    # Get all invoices for this supplier (no date filter for full history)
    query = {
        "user_id": current_user.user_id,
        "supplier": {"$regex": f"^{supplier_name}$", "$options": "i"}
    }
    
    invoices = await db.invoices.find(query, {"_id": 0, "image_base64": 0}).sort("date", 1).to_list(10000)
    
    if not invoices:
        return {
            "supplier": supplier_name,
            "found": False,
            "invoice_count": 0
        }
    
    now = datetime.now(timezone.utc)
    
    # Basic stats
    total_amount = sum(inv.get("total_amount", 0) for inv in invoices)
    total_vat = sum(inv.get("vat_amount", 0) for inv in invoices)
    total_net = sum(inv.get("amount_without_vat", 0) for inv in invoices)
    avg_invoice = total_amount / len(invoices) if invoices else 0
    
    # Extract dates
    dates = []
    amounts = []
    for inv in invoices:
        date_val = inv.get("date")
        if isinstance(date_val, datetime):
            dates.append(date_val)
            amounts.append(inv.get("total_amount", 0))
    
    first_delivery = min(dates) if dates else None
    last_delivery = max(dates) if dates else None
    
    # Ensure timezone consistency for date calculations
    if last_delivery:
        if last_delivery.tzinfo is None:
            last_delivery = last_delivery.replace(tzinfo=timezone.utc)
        days_inactive = (now - last_delivery).days
    else:
        days_inactive = 999
    
    is_active = days_inactive <= 30
    
    # Monthly breakdown
    monthly_data = defaultdict(lambda: {"amount": 0, "count": 0})
    for inv in invoices:
        date_val = inv.get("date")
        if isinstance(date_val, datetime):
            month_key = date_val.strftime("%Y-%m")
            monthly_data[month_key]["amount"] += inv.get("total_amount", 0)
            monthly_data[month_key]["count"] += 1
    
    # Sort monthly data
    sorted_months = sorted(monthly_data.keys())
    monthly_trend = []
    prev_amount = 0
    for month in sorted_months[-12:]:  # Last 12 months
        data = monthly_data[month]
        growth = ((data["amount"] - prev_amount) / prev_amount * 100) if prev_amount > 0 else 0
        monthly_trend.append({
            "month": month,
            "amount": round(data["amount"], 2),
            "count": data["count"],
            "growth_percent": round(growth, 1)
        })
        prev_amount = data["amount"]
    
    # Calculate anomalies
    if len(amounts) >= 3:
        mean_amount = sum(amounts) / len(amounts)
        variance = sum((x - mean_amount) ** 2 for x in amounts) / len(amounts)
        std_dev = variance ** 0.5
        threshold = mean_amount + (2 * std_dev)
        
        anomalies = []
        for inv in invoices:
            if inv.get("total_amount", 0) > threshold:
                date_val = inv.get("date")
                date_str = date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else str(date_val)[:10]
                anomalies.append({
                    "date": date_str,
                    "amount": inv.get("total_amount", 0),
                    "invoice_number": inv.get("invoice_number"),
                    "deviation_percent": round((inv.get("total_amount", 0) - mean_amount) / mean_amount * 100, 1)
                })
    else:
        anomalies = []
    
    # Recent invoices
    recent_invoices = []
    for inv in invoices[-10:][::-1]:
        date_val = inv.get("date")
        date_str = date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else str(date_val)[:10]
        recent_invoices.append({
            "id": inv.get("id"),
            "invoice_number": inv.get("invoice_number"),
            "date": date_str,
            "total_amount": inv.get("total_amount", 0),
            "vat_amount": inv.get("vat_amount", 0)
        })
    
    return {
        "supplier": supplier_name,
        "found": True,
        "overview": {
            "total_amount": round(total_amount, 2),
            "total_vat": round(total_vat, 2),
            "total_net": round(total_net, 2),
            "invoice_count": len(invoices),
            "avg_invoice": round(avg_invoice, 2),
            "first_delivery": first_delivery.strftime("%Y-%m-%d") if first_delivery else None,
            "last_delivery": last_delivery.strftime("%Y-%m-%d") if last_delivery else None,
            "is_active": is_active,
            "days_inactive": days_inactive if not is_active else 0
        },
        "monthly_trend": monthly_trend,
        "anomalies": anomalies,
        "recent_invoices": recent_invoices
    }

@api_router.get("/statistics/suppliers/compare")
async def compare_suppliers(
    suppliers: str,  # comma-separated supplier names
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Compare multiple suppliers"""
    from urllib.parse import unquote
    
    supplier_names = [unquote(s.strip()) for s in suppliers.split(",") if s.strip()]
    
    if len(supplier_names) < 2 or len(supplier_names) > 5:
        raise HTTPException(status_code=400, detail="Изберете между 2 и 5 доставчика за сравнение")
    
    # Build date query
    date_query = {}
    if start_date:
        date_query["$gte"] = datetime.fromisoformat(start_date + "T00:00:00+00:00")
    if end_date:
        date_query["$lte"] = datetime.fromisoformat(end_date + "T23:59:59+00:00")
    
    comparison = []
    
    for supplier_name in supplier_names:
        query = {
            "user_id": current_user.user_id,
            "supplier": {"$regex": f"^{supplier_name}$", "$options": "i"}
        }
        if date_query:
            query["date"] = date_query
        
        invoices = await db.invoices.find(query, {"_id": 0, "total_amount": 1, "date": 1}).to_list(10000)
        
        total = sum(inv.get("total_amount", 0) for inv in invoices)
        count = len(invoices)
        avg = total / count if count > 0 else 0
        
        comparison.append({
            "supplier": supplier_name,
            "total_amount": round(total, 2),
            "invoice_count": count,
            "avg_invoice": round(avg, 2)
        })
    
    # Calculate max for percentage calculation
    max_amount = max(c["total_amount"] for c in comparison) if comparison else 1
    max_count = max(c["invoice_count"] for c in comparison) if comparison else 1
    
    for c in comparison:
        c["amount_percent"] = round(c["total_amount"] / max_amount * 100, 1) if max_amount > 0 else 0
        c["count_percent"] = round(c["invoice_count"] / max_count * 100, 1) if max_count > 0 else 0
    
    return {
        "suppliers": comparison,
        "period": {"start_date": start_date, "end_date": end_date}
    }

@api_router.get("/statistics/supplier/{supplier_name}")
async def get_single_supplier_stats(
    supplier_name: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get detailed statistics for a specific supplier"""
    query = {
        "user_id": current_user.user_id,
        "supplier": {"$regex": f"^{supplier_name}$", "$options": "i"}
    }
    
    if start_date or end_date:
        query["date"] = {}
        if start_date:
            query["date"]["$gte"] = datetime.fromisoformat(start_date + "T00:00:00+00:00")
        if end_date:
            query["date"]["$lte"] = datetime.fromisoformat(end_date + "T23:59:59+00:00")
    
    invoices = await db.invoices.find(query, {"_id": 0, "image_base64": 0}).sort("date", -1).to_list(1000)
    
    if not invoices:
        return {
            "supplier": supplier_name,
            "invoice_count": 0,
            "total_amount": 0,
            "invoices": []
        }
    
    total_amount = sum(inv.get("total_amount", 0) for inv in invoices)
    total_vat = sum(inv.get("vat_amount", 0) for inv in invoices)
    total_net = sum(inv.get("amount_without_vat", 0) for inv in invoices)
    
    # Format invoices
    formatted_invoices = []
    for inv in invoices:
        date_val = inv.get("date")
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)[:10]
        
        formatted_invoices.append({
            "id": inv.get("id"),
            "invoice_number": inv.get("invoice_number"),
            "date": date_str,
            "total_amount": inv.get("total_amount", 0),
            "vat_amount": inv.get("vat_amount", 0),
            "amount_without_vat": inv.get("amount_without_vat", 0)
        })
    
    return {
        "supplier": supplier_name,
        "invoice_count": len(invoices),
        "total_amount": round(total_amount, 2),
        "total_vat": round(total_vat, 2),
        "total_net": round(total_net, 2),
        "avg_invoice": round(total_amount / len(invoices), 2) if invoices else 0,
        "invoices": formatted_invoices
    }

# ===================== EXPORT ENDPOINTS =====================

@api_router.get("/export/excel")
async def export_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Get data
    query = {"user_id": current_user.user_id}
    if start_date or end_date:
        query["date"] = {}
        if start_date:
            query["date"]["$gte"] = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
        if end_date:
            query["date"]["$lte"] = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
    
    invoices = await db.invoices.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Фактури"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    
    # Headers
    headers = ["Дата", "Доставчик", "№ Фактура", "Без ДДС", "ДДС", "Общо", "Бележки"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, inv in enumerate(invoices, 2):
        date_val = inv["date"]
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)[:10]
        
        ws.cell(row=row, column=1, value=date_str)
        ws.cell(row=row, column=2, value=inv.get("supplier", ""))
        ws.cell(row=row, column=3, value=inv.get("invoice_number", ""))
        ws.cell(row=row, column=4, value=inv.get("amount_without_vat", 0))
        ws.cell(row=row, column=5, value=inv.get("vat_amount", 0))
        ws.cell(row=row, column=6, value=inv.get("total_amount", 0))
        ws.cell(row=row, column=7, value=inv.get("notes", ""))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 30
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fakturi.xlsx"}
    )

@api_router.get("/export/pdf")
async def export_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    # Get data
    query = {"user_id": current_user.user_id}
    if start_date or end_date:
        query["date"] = {}
        if start_date:
            query["date"]["$gte"] = datetime.fromisoformat(start_date.replace("Z", "+00:00"))
        if end_date:
            query["date"]["$lte"] = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
    
    invoices = await db.invoices.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    
    # Table data
    data = [["Data", "Dostavchik", "No Faktura", "Bez DDS", "DDS", "Obshto"]]
    
    for inv in invoices:
        date_val = inv["date"]
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)[:10]
        
        data.append([
            date_str,
            inv.get("supplier", "")[:30],
            inv.get("invoice_number", ""),
            f"{inv.get('amount_without_vat', 0):.2f}",
            f"{inv.get('vat_amount', 0):.2f}",
            f"{inv.get('total_amount', 0):.2f}"
        ])
    
    # Create table
    table = Table(data, colWidths=[80, 150, 100, 80, 80, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    
    doc.build([table])
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=fakturi.pdf"}
    )

# ===================== BACKUP ENDPOINTS =====================

class BackupData(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    company_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    invoices: List[dict] = []
    daily_revenues: List[dict] = []
    expenses: List[dict] = []
    company: Optional[dict] = None

class BackupMetadata(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    file_name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    size_bytes: int
    invoice_count: int
    revenue_count: int
    expense_count: int
    google_drive_file_id: Optional[str] = None

@api_router.post("/backup/create")
async def create_backup(current_user: User = Depends(get_current_user)):
    """Създава backup на всички данни на потребителя"""
    import json
    
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0})
    company_id = user_doc.get("company_id") if user_doc else None
    
    # Събиране на фактури
    invoices = await db.invoices.find(
        {"user_id": current_user.user_id},
        {"_id": 0}
    ).to_list(10000)
    
    # Конвертиране на datetime обекти
    for inv in invoices:
        if isinstance(inv.get("date"), datetime):
            inv["date"] = inv["date"].isoformat()
        if isinstance(inv.get("created_at"), datetime):
            inv["created_at"] = inv["created_at"].isoformat()
    
    # Събиране на дневни обороти
    revenues = await db.daily_revenue.find(
        {"user_id": current_user.user_id},
        {"_id": 0}
    ).to_list(10000)
    
    for rev in revenues:
        if isinstance(rev.get("date"), datetime):
            rev["date"] = rev["date"].isoformat()
        if isinstance(rev.get("created_at"), datetime):
            rev["created_at"] = rev["created_at"].isoformat()
    
    # Събиране на разходи
    expenses = await db.expenses.find(
        {"user_id": current_user.user_id},
        {"_id": 0}
    ).to_list(10000)
    
    for exp in expenses:
        if isinstance(exp.get("date"), datetime):
            exp["date"] = exp["date"].isoformat()
        if isinstance(exp.get("created_at"), datetime):
            exp["created_at"] = exp["created_at"].isoformat()
    
    # Събиране на данни за фирма
    company_data = None
    if company_id:
        company_data = await db.companies.find_one({"id": company_id}, {"_id": 0})
        if company_data:
            if isinstance(company_data.get("created_at"), datetime):
                company_data["created_at"] = company_data["created_at"].isoformat()
            if isinstance(company_data.get("updated_at"), datetime):
                company_data["updated_at"] = company_data["updated_at"].isoformat()
    
    backup_data = {
        "id": str(uuid.uuid4()),
        "user_id": current_user.user_id,
        "user_email": current_user.email,
        "user_name": current_user.name,
        "company_id": company_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "app_version": "1.0.0",
        "invoices": invoices,
        "daily_revenues": revenues,
        "expenses": expenses,
        "company": company_data,
        "statistics": {
            "invoice_count": len(invoices),
            "revenue_count": len(revenues),
            "expense_count": len(expenses)
        }
    }
    
    # Записване на metadata за backup
    backup_json = json.dumps(backup_data, ensure_ascii=False)
    metadata = BackupMetadata(
        user_id=current_user.user_id,
        file_name=f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        size_bytes=len(backup_json.encode('utf-8')),
        invoice_count=len(invoices),
        revenue_count=len(revenues),
        expense_count=len(expenses)
    )
    
    await db.backup_metadata.insert_one(metadata.dict())
    
    return backup_data

@api_router.get("/backup/list")
async def list_backups(current_user: User = Depends(get_current_user)):
    """Връща списък с всички backups на потребителя"""
    backups = await db.backup_metadata.find(
        {"user_id": current_user.user_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return {"backups": backups}

@api_router.post("/backup/restore")
async def restore_backup(backup_data: dict, current_user: User = Depends(get_current_user)):
    """Възстановява данни от backup"""
    
    restored_counts = {
        "invoices": 0,
        "revenues": 0,
        "expenses": 0
    }
    
    # Възстановяване на фактури
    if "invoices" in backup_data:
        for invoice in backup_data["invoices"]:
            # Проверка за дублиране
            existing = await db.invoices.find_one({
                "user_id": current_user.user_id,
                "id": invoice.get("id")
            })
            if not existing:
                invoice["user_id"] = current_user.user_id
                if isinstance(invoice.get("date"), str):
                    invoice["date"] = datetime.fromisoformat(invoice["date"].replace("Z", "+00:00"))
                if isinstance(invoice.get("created_at"), str):
                    invoice["created_at"] = datetime.fromisoformat(invoice["created_at"].replace("Z", "+00:00"))
                await db.invoices.insert_one(invoice)
                restored_counts["invoices"] += 1
    
    # Възстановяване на дневни обороти
    if "daily_revenues" in backup_data:
        for revenue in backup_data["daily_revenues"]:
            existing = await db.daily_revenue.find_one({
                "user_id": current_user.user_id,
                "id": revenue.get("id")
            })
            if not existing:
                revenue["user_id"] = current_user.user_id
                if isinstance(revenue.get("date"), str):
                    revenue["date"] = datetime.fromisoformat(revenue["date"].replace("Z", "+00:00"))
                await db.daily_revenue.insert_one(revenue)
                restored_counts["revenues"] += 1
    
    # Възстановяване на разходи
    if "expenses" in backup_data:
        for expense in backup_data["expenses"]:
            existing = await db.expenses.find_one({
                "user_id": current_user.user_id,
                "id": expense.get("id")
            })
            if not existing:
                expense["user_id"] = current_user.user_id
                if isinstance(expense.get("date"), str):
                    expense["date"] = datetime.fromisoformat(expense["date"].replace("Z", "+00:00"))
                await db.expenses.insert_one(expense)
                restored_counts["expenses"] += 1
    
    return {
        "success": True,
        "message": "Данните са възстановени успешно",
        "restored": restored_counts
    }

@api_router.get("/backup/status")
async def get_backup_status(current_user: User = Depends(get_current_user)):
    """Връща статус на последния backup"""
    last_backup = await db.backup_metadata.find_one(
        {"user_id": current_user.user_id},
        {"_id": 0}
    )
    
    if last_backup:
        return {
            "has_backup": True,
            "last_backup_date": last_backup.get("created_at"),
            "file_name": last_backup.get("file_name"),
            "statistics": {
                "invoices": last_backup.get("invoice_count", 0),
                "revenues": last_backup.get("revenue_count", 0),
                "expenses": last_backup.get("expense_count", 0)
            }
        }
    
    return {
        "has_backup": False,
        "last_backup_date": None
    }

# ===================== ITEM PRICE TRACKING =====================

@api_router.get("/items/price-alerts")
async def get_price_alerts(
    status: Optional[str] = None,  # unread, read, dismissed
    current_user: User = Depends(get_current_user)
):
    """Връща ценови аларми за фирмата"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {"alerts": [], "total": 0, "unread_count": 0}
    
    query = {"company_id": company_id}
    if status:
        query["status"] = status
    
    alerts = await db.price_alerts.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Count unread
    unread_count = await db.price_alerts.count_documents({"company_id": company_id, "status": "unread"})
    
    return {
        "alerts": alerts,
        "total": len(alerts),
        "unread_count": unread_count
    }

@api_router.put("/items/price-alerts/{alert_id}")
async def update_price_alert(
    alert_id: str,
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Обновява статус на аларма (read, dismissed)"""
    body = await request.json()
    status = body.get("status")
    
    if status not in ["read", "dismissed"]:
        raise HTTPException(status_code=400, detail="Невалиден статус")
    
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    result = await db.price_alerts.update_one(
        {"id": alert_id, "company_id": company_id},
        {"$set": {"status": status}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Алармата не е намерена")
    
    return {"message": "Статусът е обновен"}

@api_router.get("/items/price-alert-settings")
async def get_price_alert_settings(current_user: User = Depends(get_current_user)):
    """Връща настройки за ценови аларми"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {"threshold_percent": 10.0, "enabled": True}
    
    settings = await db.price_alert_settings.find_one({"company_id": company_id}, {"_id": 0})
    
    if not settings:
        return {"threshold_percent": 10.0, "enabled": True}
    
    return {
        "threshold_percent": settings.get("threshold_percent", 10.0),
        "enabled": settings.get("enabled", True)
    }

@api_router.put("/items/price-alert-settings")
async def update_price_alert_settings(
    request: Request,
    current_user: User = Depends(get_current_user)
):
    """Обновява настройки за ценови аларми"""
    body = await request.json()
    
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        raise HTTPException(status_code=400, detail="Нямате фирма")
    
    update_data = {}
    if "threshold_percent" in body:
        update_data["threshold_percent"] = float(body["threshold_percent"])
    if "enabled" in body:
        update_data["enabled"] = bool(body["enabled"])
    
    existing = await db.price_alert_settings.find_one({"company_id": company_id})
    
    if existing:
        await db.price_alert_settings.update_one(
            {"company_id": company_id},
            {"$set": update_data}
        )
    else:
        settings = PriceAlertSettings(company_id=company_id, **update_data)
        await db.price_alert_settings.insert_one(settings.dict())
    
    return {"message": "Настройките са запазени"}

@api_router.get("/items/price-history/{item_name}")
async def get_item_price_history(
    item_name: str,
    supplier: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Връща история на цените за артикул"""
    from urllib.parse import unquote
    
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {"history": [], "statistics": {}}
    
    item_name = unquote(item_name).strip().lower()
    
    query = {
        "company_id": company_id,
        "item_name": item_name
    }
    if supplier:
        query["supplier"] = {"$regex": f"^{unquote(supplier)}$", "$options": "i"}
    
    history = await db.item_price_history.find(query, {"_id": 0}).sort("invoice_date", 1).to_list(1000)
    
    if not history:
        return {"history": [], "statistics": {}}
    
    # Calculate statistics
    prices = [h["unit_price"] for h in history]
    avg_price = sum(prices) / len(prices) if prices else 0
    min_price = min(prices) if prices else 0
    max_price = max(prices) if prices else 0
    
    # Calculate variance
    if len(prices) > 1:
        variance = sum((p - avg_price) ** 2 for p in prices) / len(prices)
        std_dev = variance ** 0.5
    else:
        std_dev = 0
    
    # Trend (last 3 vs first 3)
    if len(prices) >= 6:
        first_avg = sum(prices[:3]) / 3
        last_avg = sum(prices[-3:]) / 3
        trend_percent = ((last_avg - first_avg) / first_avg * 100) if first_avg > 0 else 0
    else:
        trend_percent = 0
    
    # Format history for response
    formatted_history = []
    for h in history:
        date_val = h.get("invoice_date")
        date_str = date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else str(date_val)[:10]
        formatted_history.append({
            "date": date_str,
            "supplier": h.get("supplier"),
            "unit_price": h.get("unit_price"),
            "quantity": h.get("quantity"),
            "unit": h.get("unit"),
            "invoice_number": h.get("invoice_number")
        })
    
    return {
        "item_name": item_name,
        "history": formatted_history,
        "statistics": {
            "avg_price": round(avg_price, 2),
            "min_price": round(min_price, 2),
            "max_price": round(max_price, 2),
            "std_dev": round(std_dev, 2),
            "trend_percent": round(trend_percent, 1),
            "total_records": len(history)
        }
    }

@api_router.get("/statistics/items")
async def get_item_statistics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    top_n: int = 10,
    current_user: User = Depends(get_current_user)
):
    """Връща статистика за артикули - топ N по брой и стойност"""
    from collections import defaultdict
    
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {
            "totals": {"total_items": 0, "total_value": 0, "unique_items": 0},
            "top_by_quantity": [],
            "top_by_value": [],
            "top_by_frequency": [],
            "price_trends": []
        }
    
    # Build date query
    query = {"company_id": company_id}
    if start_date or end_date:
        query["invoice_date"] = {}
        if start_date:
            query["invoice_date"]["$gte"] = datetime.fromisoformat(start_date + "T00:00:00+00:00")
        if end_date:
            query["invoice_date"]["$lte"] = datetime.fromisoformat(end_date + "T23:59:59+00:00")
    
    # Get all price history records
    history = await db.item_price_history.find(query, {"_id": 0}).to_list(10000)
    
    if not history:
        return {
            "totals": {"total_items": 0, "total_value": 0, "unique_items": 0},
            "top_by_quantity": [],
            "top_by_value": [],
            "top_by_frequency": [],
            "price_trends": []
        }
    
    # Aggregate by item
    item_stats = defaultdict(lambda: {
        "quantity": 0,
        "total_value": 0,
        "frequency": 0,
        "prices": [],
        "suppliers": set()
    })
    
    for record in history:
        name = record["item_name"]
        price = record["unit_price"]
        qty = record["quantity"]
        
        item_stats[name]["quantity"] += qty
        item_stats[name]["total_value"] += price * qty
        item_stats[name]["frequency"] += 1
        item_stats[name]["prices"].append(price)
        item_stats[name]["suppliers"].add(record["supplier"])
    
    # Calculate statistics for each item
    items_list = []
    for name, stats in item_stats.items():
        prices = stats["prices"]
        avg_price = sum(prices) / len(prices) if prices else 0
        
        # Price variance
        if len(prices) > 1:
            variance = sum((p - avg_price) ** 2 for p in prices) / len(prices)
            price_variance = (variance ** 0.5) / avg_price * 100 if avg_price > 0 else 0
        else:
            price_variance = 0
        
        # Trend
        if len(prices) >= 4:
            first_half = sum(prices[:len(prices)//2]) / (len(prices)//2)
            second_half = sum(prices[len(prices)//2:]) / (len(prices) - len(prices)//2)
            trend = ((second_half - first_half) / first_half * 100) if first_half > 0 else 0
        else:
            trend = 0
        
        items_list.append({
            "item_name": name,
            "quantity": round(stats["quantity"], 2),
            "total_value": round(stats["total_value"], 2),
            "frequency": stats["frequency"],
            "avg_price": round(avg_price, 2),
            "min_price": round(min(prices), 2) if prices else 0,
            "max_price": round(max(prices), 2) if prices else 0,
            "price_variance": round(price_variance, 1),
            "trend_percent": round(trend, 1),
            "supplier_count": len(stats["suppliers"])
        })
    
    # Sort by different criteria
    top_by_quantity = sorted(items_list, key=lambda x: x["quantity"], reverse=True)[:top_n]
    top_by_value = sorted(items_list, key=lambda x: x["total_value"], reverse=True)[:top_n]
    top_by_frequency = sorted(items_list, key=lambda x: x["frequency"], reverse=True)[:top_n]
    
    # Items with significant price changes
    price_trends = sorted(
        [i for i in items_list if abs(i["trend_percent"]) > 5],
        key=lambda x: abs(x["trend_percent"]),
        reverse=True
    )[:top_n]
    
    # Totals
    total_quantity = sum(i["quantity"] for i in items_list)
    total_value = sum(i["total_value"] for i in items_list)
    
    return {
        "totals": {
            "total_items": round(total_quantity, 2),
            "total_value": round(total_value, 2),
            "unique_items": len(items_list)
        },
        "top_by_quantity": top_by_quantity,
        "top_by_value": top_by_value,
        "top_by_frequency": top_by_frequency,
        "price_trends": price_trends
    }

@api_router.get("/statistics/items/{item_name}/by-supplier")
async def get_item_by_supplier(
    item_name: str,
    current_user: User = Depends(get_current_user)
):
    """Сравнява цените на артикул между различни доставчици"""
    from collections import defaultdict
    from urllib.parse import unquote
    
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {"item_name": item_name, "suppliers": []}
    
    normalized_name = unquote(item_name).strip().lower()
    
    history = await db.item_price_history.find(
        {"company_id": company_id, "item_name": normalized_name},
        {"_id": 0}
    ).to_list(10000)
    
    if not history:
        return {"item_name": item_name, "suppliers": [], "recommendation": None}
    
    # Group by supplier
    supplier_data = defaultdict(lambda: {"prices": [], "quantities": [], "dates": []})
    
    for record in history:
        supplier = record["supplier"]
        supplier_data[supplier]["prices"].append(record["unit_price"])
        supplier_data[supplier]["quantities"].append(record["quantity"])
        date_val = record.get("invoice_date")
        if isinstance(date_val, datetime):
            supplier_data[supplier]["dates"].append(date_val)
    
    # Calculate stats per supplier
    suppliers = []
    for supplier, data in supplier_data.items():
        prices = data["prices"]
        avg = sum(prices) / len(prices) if prices else 0
        last_date = max(data["dates"]) if data["dates"] else None
        
        suppliers.append({
            "supplier": supplier,
            "avg_price": round(avg, 2),
            "min_price": round(min(prices), 2) if prices else 0,
            "max_price": round(max(prices), 2) if prices else 0,
            "last_price": round(prices[-1], 2) if prices else 0,
            "purchase_count": len(prices),
            "total_quantity": round(sum(data["quantities"]), 2),
            "last_purchase": last_date.strftime("%Y-%m-%d") if last_date else None
        })
    
    # Sort by average price (cheapest first)
    suppliers.sort(key=lambda x: x["avg_price"])
    
    # Recommendation
    recommendation = None
    if len(suppliers) >= 2:
        cheapest = suppliers[0]
        most_expensive = suppliers[-1]
        if cheapest["avg_price"] > 0:
            savings_percent = ((most_expensive["avg_price"] - cheapest["avg_price"]) / most_expensive["avg_price"] * 100)
            if savings_percent > 5:
                recommendation = {
                    "best_supplier": cheapest["supplier"],
                    "avg_price": cheapest["avg_price"],
                    "potential_savings_percent": round(savings_percent, 1)
                }
    
    return {
        "item_name": item_name,
        "suppliers": suppliers,
        "recommendation": recommendation
    }

# ===================== AI ITEM MERGING =====================

@api_router.post("/items/ai-merge")
async def ai_merge_similar_items(
    current_user: User = Depends(get_current_user)
):
    """
    AI модул за автоматично сливане на сходни продукти.
    Използва Gemini за идентифициране на еднакви продукти с различни имена.
    """
    from collections import defaultdict
    
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {"merged_groups": [], "total_merged": 0}
    
    # Get all unique item names
    history = await db.item_price_history.find(
        {"company_id": company_id},
        {"_id": 0, "item_name": 1}
    ).to_list(10000)
    
    unique_items = list(set(h["item_name"] for h in history))
    
    if len(unique_items) < 2:
        return {"merged_groups": [], "total_merged": 0, "message": "Недостатъчно артикули за анализ"}
    
    # Use AI to find similar items
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        items_text = "\n".join(unique_items[:100])  # Limit to 100 items
        
        prompt = f"""Анализирай следния списък с имена на продукти и групирай сходните продукти.
Търси продукти, които са едни и същи, но са записани по различен начин (различен регистър, съкращения, правописни грешки, варианти на името).

Списък с продукти:
{items_text}

Върни САМО JSON масив с групите, без допълнителен текст. Формат:
[
  {{"canonical_name": "Олио слънчогледово", "variants": ["Олио Първа Преса", "олио слънчогледово", "ОЛИО", "Олио Екстра"]}},
  {{"canonical_name": "Захар кристална", "variants": ["Захар", "захар кристална", "ЗАХАР БГ"]}}
]

Ако няма сходни продукти, върни празен масив: []
Не групирай продукти, които са наистина различни (например "Олио" и "Оцет" са РАЗЛИЧНИ).
Групирай САМО ако са очевидно същият продукт с различно изписване."""

        llm = LlmChat(
            api_key=os.getenv("EMERGENT_LLM_KEY"),
            session_id=f"merge_{uuid.uuid4().hex[:8]}",
            system_message="Ти си асистент за анализ на продукти. Отговаряй само с валиден JSON."
        ).with_model("gemini", "gemini-2.5-flash")
        
        user_message = UserMessage(text=prompt)
        response = await llm.send_message(user_message)
        response_text = response.strip() if isinstance(response, str) else str(response)
        
        # Extract JSON from response
        import re
        json_match = re.search(r'\[[\s\S]*\]', response_text)
        if json_match:
            merged_groups = json.loads(json_match.group())
        else:
            merged_groups = []
        
        # Save merge mappings to database
        if merged_groups:
            for group in merged_groups:
                canonical = group.get("canonical_name", "")
                variants = group.get("variants", [])
                
                if canonical and variants:
                    # Create or update merge mapping
                    await db.item_merge_mappings.update_one(
                        {"company_id": company_id, "canonical_name": canonical.lower()},
                        {
                            "$set": {
                                "canonical_name": canonical.lower(),
                                "display_name": canonical,
                                "variants": [v.lower() for v in variants],
                                "company_id": company_id,
                                "updated_at": datetime.now(timezone.utc)
                            }
                        },
                        upsert=True
                    )
        
        total_merged = sum(len(g.get("variants", [])) for g in merged_groups)
        
        return {
            "merged_groups": merged_groups,
            "total_merged": total_merged,
            "message": f"Намерени {len(merged_groups)} групи сходни продукти"
        }
        
    except Exception as e:
        logger.error(f"AI merge error: {str(e)}")
        return {"merged_groups": [], "total_merged": 0, "error": str(e)}

@api_router.get("/items/merge-mappings")
async def get_merge_mappings(current_user: User = Depends(get_current_user)):
    """Връща текущите сливания на продукти"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {"mappings": []}
    
    mappings = await db.item_merge_mappings.find(
        {"company_id": company_id},
        {"_id": 0}
    ).to_list(1000)
    
    return {"mappings": mappings}

@api_router.delete("/items/merge-mappings/{canonical_name}")
async def delete_merge_mapping(
    canonical_name: str,
    current_user: User = Depends(get_current_user)
):
    """Изтрива сливане на продукти"""
    from urllib.parse import unquote
    
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        raise HTTPException(status_code=404, detail="Фирмата не е намерена")
    
    result = await db.item_merge_mappings.delete_one({
        "company_id": company_id,
        "canonical_name": unquote(canonical_name).lower()
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Сливането не е намерено")
    
    return {"message": "Сливането е изтрито"}

@api_router.get("/statistics/items/merged")
async def get_merged_item_statistics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    top_n: int = 10,
    current_user: User = Depends(get_current_user)
):
    """
    Връща статистика за артикули с приложени AI сливания.
    Сходните продукти са обединени в една позиция.
    """
    from collections import defaultdict
    
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {
            "totals": {"total_items": 0, "total_value": 0, "unique_items": 0, "merged_items": 0},
            "top_by_quantity": [],
            "top_by_value": [],
            "merge_applied": False
        }
    
    # Get merge mappings
    mappings = await db.item_merge_mappings.find(
        {"company_id": company_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Create variant to canonical mapping
    variant_to_canonical = {}
    canonical_display = {}
    for m in mappings:
        canonical = m["canonical_name"]
        canonical_display[canonical] = m.get("display_name", canonical)
        for variant in m.get("variants", []):
            variant_to_canonical[variant.lower()] = canonical
    
    # Build date query
    query = {"company_id": company_id}
    if start_date or end_date:
        query["invoice_date"] = {}
        if start_date:
            query["invoice_date"]["$gte"] = datetime.fromisoformat(start_date + "T00:00:00+00:00")
        if end_date:
            query["invoice_date"]["$lte"] = datetime.fromisoformat(end_date + "T23:59:59+00:00")
    
    # Get all price history records
    history = await db.item_price_history.find(query, {"_id": 0}).to_list(10000)
    
    if not history:
        return {
            "totals": {"total_items": 0, "total_value": 0, "unique_items": 0, "merged_items": 0},
            "top_by_quantity": [],
            "top_by_value": [],
            "merge_applied": bool(mappings)
        }
    
    # Aggregate with merging
    item_stats = defaultdict(lambda: {
        "display_name": "",
        "quantity": 0,
        "total_value": 0,
        "frequency": 0,
        "prices": [],
        "suppliers": set(),
        "original_names": set()
    })
    
    merged_count = 0
    for record in history:
        original_name = record["item_name"]
        name_lower = original_name.lower()
        
        # Apply merge mapping
        if name_lower in variant_to_canonical:
            canonical = variant_to_canonical[name_lower]
            display = canonical_display.get(canonical, canonical.title())
            merged_count += 1
        else:
            canonical = name_lower
            display = original_name
        
        price = record["unit_price"]
        qty = record["quantity"]
        
        item_stats[canonical]["display_name"] = display
        item_stats[canonical]["quantity"] += qty
        item_stats[canonical]["total_value"] += price * qty
        item_stats[canonical]["frequency"] += 1
        item_stats[canonical]["prices"].append(price)
        item_stats[canonical]["suppliers"].add(record["supplier"])
        item_stats[canonical]["original_names"].add(original_name)
    
    # Calculate totals
    total_items = sum(s["frequency"] for s in item_stats.values())
    total_value = sum(s["total_value"] for s in item_stats.values())
    unique_items = len(item_stats)
    
    # Build top lists
    items_list = []
    for canonical, stats in item_stats.items():
        prices = stats["prices"]
        avg_price = sum(prices) / len(prices) if prices else 0
        
        items_list.append({
            "name": stats["display_name"],
            "canonical_name": canonical,
            "total_quantity": round(stats["quantity"], 2),
            "total_value": round(stats["total_value"], 2),
            "frequency": stats["frequency"],
            "avg_price": round(avg_price, 2),
            "supplier_count": len(stats["suppliers"]),
            "original_names": list(stats["original_names"])[:5]  # Show max 5 variants
        })
    
    # Sort for top lists
    top_by_quantity = sorted(items_list, key=lambda x: x["total_quantity"], reverse=True)[:top_n]
    top_by_value = sorted(items_list, key=lambda x: x["total_value"], reverse=True)[:top_n]
    
    return {
        "totals": {
            "total_items": total_items,
            "total_value": round(total_value, 2),
            "unique_items": unique_items,
            "merged_items": merged_count
        },
        "top_by_quantity": top_by_quantity,
        "top_by_value": top_by_value,
        "merge_applied": bool(mappings),
        "merge_groups_count": len(mappings)
    }

# ===================== EXPORT ENDPOINTS =====================

from services.export_service import ExportService
from services.audit_service import AuditService
from services.forecast_service import ForecastService

# Initialize services
audit_service = AuditService(db)
forecast_service = ForecastService(db)

@api_router.get("/export/invoices/excel")
async def export_invoices_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export invoices to Excel"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    query = {}
    if company_id:
        query["company_id"] = company_id
    else:
        query["user_id"] = current_user.user_id
    
    if start_date:
        query["date"] = {"$gte": datetime.fromisoformat(start_date + "T00:00:00+00:00")}
    if end_date:
        if "date" not in query:
            query["date"] = {}
        query["date"]["$lte"] = datetime.fromisoformat(end_date + "T23:59:59+00:00")
    
    invoices = await db.invoices.find(query, {"_id": 0, "image_base64": 0}).sort("date", -1).to_list(10000)
    
    # Get company name
    company_name = ""
    if company_id:
        company = await db.companies.find_one({"id": company_id})
        company_name = company.get("name", "") if company else ""
    
    try:
        excel_data = ExportService.generate_invoices_excel(invoices, company_name)
        
        # Log export
        await audit_service.log_action(
            user_id=current_user.user_id,
            user_name=current_user.name,
            action="export",
            entity_type="invoices",
            company_id=company_id,
            details={"format": "excel", "count": len(invoices)}
        )
        
        filename = f"invoices_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export not available")

@api_router.get("/export/invoices/pdf")
async def export_invoices_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export invoices to PDF"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    query = {}
    if company_id:
        query["company_id"] = company_id
    else:
        query["user_id"] = current_user.user_id
    
    if start_date:
        query["date"] = {"$gte": datetime.fromisoformat(start_date + "T00:00:00+00:00")}
    if end_date:
        if "date" not in query:
            query["date"] = {}
        query["date"]["$lte"] = datetime.fromisoformat(end_date + "T23:59:59+00:00")
    
    invoices = await db.invoices.find(query, {"_id": 0, "image_base64": 0}).sort("date", -1).to_list(10000)
    
    company_name = ""
    if company_id:
        company = await db.companies.find_one({"id": company_id})
        company_name = company.get("name", "") if company else ""
    
    try:
        pdf_data = ExportService.generate_invoices_pdf(invoices, company_name)
        
        await audit_service.log_action(
            user_id=current_user.user_id,
            user_name=current_user.name,
            action="export",
            entity_type="invoices",
            company_id=company_id,
            details={"format": "pdf", "count": len(invoices)}
        )
        
        filename = f"invoices_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        return Response(
            content=pdf_data,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="PDF export not available")

# ===================== BUDGET ENDPOINTS =====================

class BudgetCreate(BaseModel):
    month: str  # YYYY-MM
    expense_limit: float
    alert_threshold: float = 80.0

@api_router.get("/budget")
async def get_budgets(current_user: User = Depends(get_current_user)):
    """Get all budgets for company"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {"budgets": []}
    
    budgets = await db.budgets.find({"company_id": company_id}, {"_id": 0}).sort("month", -1).to_list(24)
    return {"budgets": budgets}

@api_router.post("/budget")
async def create_budget(budget: BudgetCreate, current_user: User = Depends(get_current_user)):
    """Create or update budget for a month"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        raise HTTPException(status_code=400, detail="No company associated")
    
    # Check if budget exists for this month
    existing = await db.budgets.find_one({"company_id": company_id, "month": budget.month})
    
    if existing:
        await db.budgets.update_one(
            {"company_id": company_id, "month": budget.month},
            {"$set": {"expense_limit": budget.expense_limit, "alert_threshold": budget.alert_threshold}}
        )
    else:
        budget_doc = {
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "month": budget.month,
            "expense_limit": budget.expense_limit,
            "alert_threshold": budget.alert_threshold,
            "created_at": datetime.now(timezone.utc)
        }
        await db.budgets.insert_one(budget_doc)
    
    return {"message": "Budget saved"}

@api_router.get("/budget/status")
async def get_budget_status(current_user: User = Depends(get_current_user)):
    """Get current month budget status"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {"has_budget": False}
    
    current_month = datetime.now().strftime("%Y-%m")
    budget = await db.budgets.find_one({"company_id": company_id, "month": current_month}, {"_id": 0})
    
    if not budget:
        return {"has_budget": False}
    
    # Calculate current expenses
    month_start = datetime.fromisoformat(f"{current_month}-01T00:00:00+00:00")
    
    invoices = await db.invoices.find(
        {"company_id": company_id, "date": {"$gte": month_start}},
        {"total_amount": 1}
    ).to_list(10000)
    
    expenses = await db.non_invoice_expenses.find(
        {"company_id": company_id, "date": {"$gte": month_start}},
        {"amount": 1}
    ).to_list(10000)
    
    total_spent = sum(inv.get("total_amount", 0) for inv in invoices)
    total_spent += sum(exp.get("amount", 0) for exp in expenses)
    
    limit = budget.get("expense_limit", 0)
    threshold = budget.get("alert_threshold", 80)
    
    percent_used = (total_spent / limit * 100) if limit > 0 else 0
    is_alert = percent_used >= threshold
    is_exceeded = percent_used >= 100
    
    return {
        "has_budget": True,
        "month": current_month,
        "expense_limit": limit,
        "total_spent": round(total_spent, 2),
        "remaining": round(max(0, limit - total_spent), 2),
        "percent_used": round(percent_used, 1),
        "alert_threshold": threshold,
        "is_alert": is_alert,
        "is_exceeded": is_exceeded
    }

# ===================== RECURRING EXPENSES =====================

class RecurringExpenseCreate(BaseModel):
    description: str
    amount: float
    day_of_month: int  # 1-28
    category: Optional[str] = None

@api_router.get("/recurring-expenses")
async def get_recurring_expenses(current_user: User = Depends(get_current_user)):
    """Get recurring expenses"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {"recurring_expenses": []}
    
    expenses = await db.recurring_expenses.find(
        {"company_id": company_id},
        {"_id": 0}
    ).to_list(100)
    
    return {"recurring_expenses": expenses}

@api_router.post("/recurring-expenses")
async def create_recurring_expense(
    expense: RecurringExpenseCreate,
    current_user: User = Depends(get_current_user)
):
    """Create recurring expense"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        raise HTTPException(status_code=400, detail="No company associated")
    
    if expense.day_of_month < 1 or expense.day_of_month > 28:
        raise HTTPException(status_code=400, detail="Day must be between 1 and 28")
    
    expense_doc = {
        "id": str(uuid.uuid4()),
        "company_id": company_id,
        "user_id": current_user.user_id,
        "description": expense.description,
        "amount": expense.amount,
        "day_of_month": expense.day_of_month,
        "category": expense.category,
        "is_active": True,
        "last_generated": None,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.recurring_expenses.insert_one(expense_doc)
    return {"message": "Recurring expense created", "id": expense_doc["id"]}

@api_router.delete("/recurring-expenses/{expense_id}")
async def delete_recurring_expense(expense_id: str, current_user: User = Depends(get_current_user)):
    """Delete recurring expense"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    result = await db.recurring_expenses.delete_one({"id": expense_id, "company_id": company_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    return {"message": "Deleted"}

# ===================== FORECAST ENDPOINTS =====================

@api_router.get("/forecast/expenses")
async def get_expense_forecast(
    months_ahead: int = 3,
    current_user: User = Depends(get_current_user)
):
    """Get expense forecast"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {"error": "No company"}
    
    return await forecast_service.get_expense_forecast(company_id, months_ahead)

@api_router.get("/forecast/revenue")
async def get_revenue_forecast(
    months_ahead: int = 3,
    current_user: User = Depends(get_current_user)
):
    """Get revenue forecast"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    
    if not company_id:
        return {"error": "No company"}
    
    return await forecast_service.get_revenue_forecast(company_id, months_ahead)

# ===================== AUDIT LOG =====================

@api_router.get("/audit-logs")
async def get_audit_logs(
    action: Optional[str] = None,
    entity_type: Optional[str] = None,
    limit: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get audit logs (Owner/Manager only)"""
    user_doc = await db.users.find_one({"user_id": current_user.user_id}, {"_id": 0, "company_id": 1, "role": 1})
    company_id = user_doc.get("company_id") if user_doc else None
    role = user_doc.get("role", "staff") if user_doc else "staff"
    
    if role not in ["owner", "manager"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    logs = await audit_service.get_logs(
        company_id=company_id,
        action=action,
        entity_type=entity_type,
        limit=limit
    )
    
    return {"logs": logs}

# ===================== DATABASE INDEXES =====================

@app.on_event("startup")
async def create_indexes():
    """Create database indexes for performance"""
    try:
        # Invoices indexes
        await db.invoices.create_index([("company_id", 1), ("date", -1)])
        await db.invoices.create_index([("company_id", 1), ("supplier", 1)])
        await db.invoices.create_index([("user_id", 1), ("date", -1)])
        
        # Revenues indexes
        await db.daily_revenues.create_index([("company_id", 1), ("date", -1)])
        
        # Expenses indexes
        await db.non_invoice_expenses.create_index([("company_id", 1), ("date", -1)])
        
        # Price history indexes
        await db.item_price_history.create_index([("company_id", 1), ("item_name", 1)])
        await db.item_price_history.create_index([("company_id", 1), ("supplier", 1)])
        
        # Users indexes
        await db.users.create_index([("email", 1)], unique=True, sparse=True)
        await db.users.create_index([("company_id", 1)])
        
        # Audit log index
        await db.audit_logs.create_index([("company_id", 1), ("created_at", -1)])
        
        logger.info("Database indexes created successfully")
    except Exception as e:
        logger.error(f"Error creating indexes: {e}")

# ===================== HEALTH CHECK =====================

@api_router.get("/")
async def root():
    return {"message": "Invoice Manager API", "status": "running"}

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

# Include the router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
