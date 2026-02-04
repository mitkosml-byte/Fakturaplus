"""Export service for generating Excel and PDF reports"""
import io
from datetime import datetime
from typing import List, Optional
import json

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class ExportService:
    @staticmethod
    def generate_invoices_excel(invoices: List[dict], company_name: str = "") -> bytes:
        """Generate Excel file from invoices"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is not installed")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Фактури"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Справка за фактури - {company_name}" if company_name else "Справка за фактури"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Генерирано на: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ["Дата", "Доставчик", "№ Фактура", "Сума без ДДС", "ДДС", "Обща сума", "Бележки"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        total_without_vat = 0
        total_vat = 0
        total_amount = 0
        
        for row, inv in enumerate(invoices, 5):
            date_val = inv.get('date', '')
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%d.%m.%Y')
            else:
                date_str = str(date_val)[:10] if date_val else ''
            
            amount_without_vat = float(inv.get('amount_without_vat', 0))
            vat_amount = float(inv.get('vat_amount', 0))
            total = float(inv.get('total_amount', 0))
            
            total_without_vat += amount_without_vat
            total_vat += vat_amount
            total_amount += total
            
            data = [
                date_str,
                inv.get('supplier', ''),
                inv.get('invoice_number', ''),
                amount_without_vat,
                vat_amount,
                total,
                inv.get('notes', '') or ''
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col >= 4 and col <= 6:
                    cell.number_format = '#,##0.00 €'
                    cell.alignment = Alignment(horizontal="right")
        
        # Totals row
        total_row = len(invoices) + 5
        ws.cell(row=total_row, column=3, value="ОБЩО:").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_without_vat).font = Font(bold=True)
        ws.cell(row=total_row, column=4).number_format = '#,##0.00 €'
        ws.cell(row=total_row, column=5, value=total_vat).font = Font(bold=True)
        ws.cell(row=total_row, column=5).number_format = '#,##0.00 €'
        ws.cell(row=total_row, column=6, value=total_amount).font = Font(bold=True)
        ws.cell(row=total_row, column=6).number_format = '#,##0.00 €'
        
        # Column widths
        column_widths = [12, 30, 15, 15, 12, 15, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def generate_invoices_pdf(invoices: List[dict], company_name: str = "") -> bytes:
        """Generate PDF file from invoices"""
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is not installed")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=10,
            alignment=1  # Center
        )
        title = f"Справка за фактури - {company_name}" if company_name else "Справка за фактури"
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"Генерирано: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 10*mm))
        
        # Table data
        data = [["Дата", "Доставчик", "No Фактура", "Без ДДС", "ДДС", "Общо"]]
        
        total_without_vat = 0
        total_vat = 0
        total_amount = 0
        
        for inv in invoices:
            date_val = inv.get('date', '')
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%d.%m.%Y')
            else:
                date_str = str(date_val)[:10] if date_val else ''
            
            amount_without_vat = float(inv.get('amount_without_vat', 0))
            vat_amount = float(inv.get('vat_amount', 0))
            total = float(inv.get('total_amount', 0))
            
            total_without_vat += amount_without_vat
            total_vat += vat_amount
            total_amount += total
            
            supplier = inv.get('supplier', '')[:20]  # Truncate long names
            
            data.append([
                date_str,
                supplier,
                inv.get('invoice_number', ''),
                f"{amount_without_vat:.2f}",
                f"{vat_amount:.2f}",
                f"{total:.2f}"
            ])
        
        # Totals
        data.append(["", "", "ОБЩО:", f"{total_without_vat:.2f}", f"{total_vat:.2f}", f"{total_amount:.2f}"])
        
        # Create table
        table = Table(data, colWidths=[55, 100, 70, 55, 45, 55])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#F8FAFC')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E2E8F0')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ]))
        
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def generate_statistics_excel(stats: dict, company_name: str = "") -> bytes:
        """Generate Excel statistics report"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is not installed")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Статистика"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = f"Финансова статистика - {company_name}" if company_name else "Финансова статистика"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A2'] = f"Генерирано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        # Summary section
        row = 4
        ws.cell(row=row, column=1, value="Показател").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value="Стойност").font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        
        summary_data = [
            ("Общо приходи", stats.get('total_income', 0)),
            ("Общо разходи", stats.get('total_expense', 0)),
            ("Печалба", stats.get('profit', 0)),
            ("ДДС за плащане", stats.get('vat_to_pay', 0)),
            ("Брой фактури", stats.get('invoice_count', 0)),
        ]
        
        for i, (label, value) in enumerate(summary_data, row + 1):
            ws.cell(row=i, column=1, value=label)
            cell = ws.cell(row=i, column=2, value=value)
            if isinstance(value, (int, float)) and label != "Брой фактури":
                cell.number_format = '#,##0.00 €'
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
